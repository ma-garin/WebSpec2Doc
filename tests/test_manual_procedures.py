"""第3弾 S2: 文書駆動パスから生成する手動テスト手順書の公開契約。"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from mbt.manual_procedures import build_manual_procedures, save_manual_procedures


def test_build_manual_procedure_keeps_requirement_and_measured_screenshot_trace() -> None:
    model = {
        "paths": [
            {
                "path_id": "DMBT-001",
                "node_ids": ["P001", "P002"],
                "requirement_ids": ["REQ-01"],
                "review_required": True,
            }
        ]
    }
    report = {
        "screens": [
            {
                "page_id": "P001",
                "title": "入口",
                "url": "https://example.com/",
                "screenshot_path": "screens/P001.png",
            },
            {
                "page_id": "P002",
                "title": "検索",
                "url": "https://example.com/search",
                "screenshot_path": "screens/P002.png",
            },
        ]
    }

    procedures = build_manual_procedures(model, report)

    assert procedures == [
        {
            "case_id": "DMBT-001",
            "title": "入口 → 検索",
            "requirement_ids": ["REQ-01"],
            "preconditions": ["対象サイトへアクセスできること"],
            "steps": [
                {
                    "step_no": 1,
                    "action": "https://example.com/ を開く",
                    "expected_result": "「入口」が表示される（生成結果・レビュー必須）",
                    "page_id": "P001",
                    "screenshot_path": "screens/P001.png",
                },
                {
                    "step_no": 2,
                    "action": "「入口」から「検索」へ遷移する",
                    "expected_result": "「検索」が表示される（生成結果・レビュー必須）",
                    "page_id": "P002",
                    "screenshot_path": "screens/P002.png",
                },
            ],
            "evidence": "measured_path",
            "review_required": True,
        }
    ]


def test_save_manual_procedures_writes_reviewable_markdown_and_excel(tmp_path: Path) -> None:
    procedures = [
        {
            "case_id": "DMBT-001",
            "title": "入口 → 検索",
            "requirement_ids": ["REQ-01"],
            "preconditions": ["対象サイトへアクセスできること"],
            "steps": [
                {
                    "step_no": 1,
                    "action": "https://example.com/ を開く",
                    "expected_result": "「入口」が表示される（生成結果・レビュー必須）",
                    "page_id": "P001",
                    "screenshot_path": "screens/P001.png",
                }
            ],
            "evidence": "measured_path",
            "review_required": True,
        }
    ]

    paths = save_manual_procedures(procedures, tmp_path)

    markdown = paths["manual_procedures_md"].read_text(encoding="utf-8")
    assert "期待結果は生成物であり、実行前レビューが必要です" in markdown
    assert "REQ-01" in markdown
    assert "![P001](screens/P001.png)" in markdown
    workbook = load_workbook(paths["manual_procedures_xlsx"])
    sheet = workbook["手動テスト手順"]
    assert list(next(sheet.iter_rows(values_only=True))) == [
        "ケースID",
        "要件ID",
        "タイトル",
        "ステップ",
        "操作",
        "期待結果（レビュー必須）",
        "画面ID",
        "スクリーンショット",
        "根拠",
    ]
    assert list(next(sheet.iter_rows(min_row=2, values_only=True)))[:4] == [
        "DMBT-001",
        "REQ-01",
        "入口 → 検索",
        1,
    ]
    assert sheet["H2"].hyperlink.target == "screens/P001.png"


def test_excel_treats_document_identifiers_as_text_not_formulas(tmp_path: Path) -> None:
    procedures = [
        {
            "case_id": "DMBT-001",
            "title": "安全確認",
            "requirement_ids": ['=HYPERLINK("https://evil.invalid")'],
            "preconditions": [],
            "steps": [
                {
                    "step_no": 1,
                    "action": "対象を確認する",
                    "expected_result": "レビュー必須",
                    "page_id": "P001",
                    "screenshot_path": "",
                }
            ],
            "evidence": "measured_path",
        }
    ]

    paths = save_manual_procedures(procedures, tmp_path)

    workbook = load_workbook(paths["manual_procedures_xlsx"], data_only=False)
    requirement_cell = workbook["手動テスト手順"]["B2"]
    assert requirement_cell.data_type == "s"
    assert requirement_cell.value.startswith("'=")
