"""Excel（.xlsx/.xlsm）からの表抽出。

シートごとにヘッダ行を自動検出し、フォーマット非依存の ExtractedTable に
変換する。ヘッダ行は既知の列名シノニムへの一致で判定するため、
表の開始位置が先頭行でなくてもよい。
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from ingest.tables import ExtractedRow, ExtractedTable, looks_like_header

_HEADER_SCAN_ROWS = 15
_EMPTY_ROW_LIMIT = 20


def read_excel_tables(path: Path) -> list[ExtractedTable]:
    """ワークブック内の全シートから表を抽出する。"""
    workbook = load_workbook(path, read_only=True, data_only=True)
    tables: list[ExtractedTable] = []
    try:
        for sheet in workbook.worksheets:
            table = _read_sheet_table(path.name, sheet)
            if table is not None:
                tables.append(table)
    finally:
        workbook.close()
    return tables


def _read_sheet_table(file_name: str, sheet: object) -> ExtractedTable | None:
    rows_iter = getattr(sheet, "iter_rows")(values_only=True)  # noqa: B009
    sheet_title = str(getattr(sheet, "title", ""))
    header: tuple[str, ...] | None = None
    header_row_number = 0
    data_rows: list[ExtractedRow] = []
    empty_streak = 0
    for row_number, raw_row in enumerate(rows_iter, start=1):
        cells = tuple("" if value is None else str(value).strip() for value in raw_row)
        if header is None:
            if row_number > _HEADER_SCAN_ROWS:
                return None
            if looks_like_header(cells):
                header = cells
                header_row_number = row_number
            continue
        if not any(cells):
            empty_streak += 1
            if empty_streak >= _EMPTY_ROW_LIMIT:
                break
            continue
        empty_streak = 0
        data_rows.append(ExtractedRow(location=f"{sheet_title}!R{row_number}", cells=cells))
    if header is None or not data_rows:
        return None
    # ヘッダ行番号は location 形式の説明用に使わないが、将来の quote 用に残す
    _ = header_row_number
    return ExtractedTable(source_file=file_name, headers=header, rows=tuple(data_rows))
