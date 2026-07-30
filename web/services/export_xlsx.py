"""テスト仕様書一式を 1 つの Excel ブックにまとめる（P2-3）。

QA 実務の納品物は Excel が主流だが、これまで `spec.xlsx` には**実測仕様の 4 シート**
（画面一覧・フォーム・項目定義書・境界値データ）しか入っておらず、
**テスト設計・テストケース・遷移表は画面でしか見られなかった**。ここで足す。

なぜ生成時ではなくエクスポート時に組み立てるか:
    テストケースはクロールより後に生成・編集され、実行結果も後から付く。
    クロール時点でブックを固めると、開いた瞬間に古い内容を渡すことになる。
    そのため要求された時点で組み直し、同じ内容をディスクにも残して
    ZIP 一括ダウンロードや CLI から読む `spec.xlsx` と食い違わないようにする。

シートの並びは画面のタブ順（読む → 設計する → 実行する）に合わせる。
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

XLSX_FILE_NAME = "spec.xlsx"

# 追加する 3 シート。既存の 4 シートの後ろに、この順で並べる。
SHEET_TEST_DESIGN = "テスト設計"
SHEET_TESTCASES = "テストケース"
SHEET_STATE_TABLE = "遷移表"
ADDED_SHEETS = (SHEET_TEST_DESIGN, SHEET_TESTCASES, SHEET_STATE_TABLE)

_HEAD_FILL = PatternFill("solid", fgColor="E3F2FD")
_HEAD_FONT = Font(bold=True)
_WRAP = Alignment(vertical="top", wrap_text=True)


class ExportError(Exception):
    """成果物が足りずブックを組み立てられない。"""


def _load_report(domain: str, out_dir: Path) -> dict[str, Any]:
    path = out_dir / domain / "report.json"
    if not path.is_file():
        raise ExportError(f"report.json がありません: {path}")
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise ExportError(f"report.json を読めません: {exc}") from exc
    if not isinstance(data, dict):
        raise ExportError("report.json の形式が想定と違います")
    return data


def _style_header(ws: openpyxl.worksheet.worksheet.Worksheet, widths: list[int]) -> None:
    for cell in ws[1]:
        cell.fill = _HEAD_FILL
        cell.font = _HEAD_FONT
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    ws.freeze_panes = "A2"


def _wrap_columns(
    ws: openpyxl.worksheet.worksheet.Worksheet, row_count: int, columns: list[int]
) -> None:
    """セル内改行を折り返して表示させる（1 始まりの列番号を渡す）。"""
    for r in range(2, row_count + 2):
        for c in columns:
            ws.cell(row=r, column=c).alignment = _WRAP


def _cell_lines(value: Any) -> str:
    """画面では複数行のリストで見えるものを、セル内改行で 1 セルにまとめる。"""
    if isinstance(value, list | tuple):
        return "\n".join(str(x) for x in value if str(x).strip())
    return "" if value is None else str(value)


# ────────────────────────── テスト設計（画面別） ──────────────────────────


def build_test_design_rows(report: dict[str, Any]) -> list[list[str]]:
    """画面別設計タブと同じ条件一覧。

    テストケースの中身（入力値・手順・期待結果）はここには載せない。
    載せるとテストケースシートと二重になり、どちらが正かが分からなくなる。
    """
    from web.services.screen_test_design import build_screen_detail, build_screen_index

    rows: list[list[str]] = []
    for screen in build_screen_index(report) or []:
        page_id = str(screen.get("page_id") or "")
        detail = build_screen_detail(report, page_id)
        if not detail:
            continue
        title = str(detail.get("title") or "")
        for cond in detail.get("conditions") or []:
            rows.append(
                [
                    page_id,
                    title,
                    str(cond.get("no") or ""),
                    str(cond.get("condition") or ""),
                    str(cond.get("technique") or ""),
                    str(cond.get("source_kind") or ""),
                    str(cond.get("source_name") or ""),
                ]
            )
    return rows


def _write_test_design(ws: openpyxl.worksheet.worksheet.Worksheet, report: dict[str, Any]) -> int:
    ws.append(["画面ID", "画面名", "No", "テスト条件", "導出技法", "由来", "由来の詳細"])
    rows = build_test_design_rows(report)
    for row in rows:
        ws.append(row)
    _style_header(ws, [10, 26, 6, 52, 14, 14, 30])
    _wrap_columns(ws, len(rows), [4])  # テスト条件の列
    return len(rows)


# ────────────────────────────── テストケース ──────────────────────────────

_CASE_COLUMNS: tuple[tuple[str, str], ...] = (
    ("case_id", "ID"),
    ("name", "テストケース名"),
    ("screen", "画面"),
    ("function", "機能"),
    ("viewpoint", "観点"),
    ("preconditions", "前提条件"),
    ("steps", "手順"),
    ("expected", "期待結果"),
    ("automation", "自動化判定"),
    ("result", "結果"),
)


def build_testcase_rows(domain: str, report: dict[str, Any]) -> list[list[str]]:
    """テストケースタブと同じ 10 列。手順・期待結果はセル内改行でまとめる。"""
    from web.services.testcase_table_store import compose

    table = compose(domain, report) or {}
    rows: list[list[str]] = []
    for row in table.get("rows") or []:
        rows.append([_cell_lines(row.get(key)) for key, _ in _CASE_COLUMNS])
    return rows


def _write_testcases(
    ws: openpyxl.worksheet.worksheet.Worksheet, domain: str, report: dict[str, Any]
) -> int:
    ws.append([label for _, label in _CASE_COLUMNS])
    rows = build_testcase_rows(domain, report)
    for row in rows:
        ws.append(row)
    _style_header(ws, [18, 34, 20, 14, 14, 28, 40, 40, 12, 10])
    _wrap_columns(ws, len(rows), [6, 7, 8])  # 前提条件・手順・期待結果
    return len(rows)


# ─────────────────────────────── 遷移表 ───────────────────────────────


def build_state_table_grid(report: dict[str, Any]) -> tuple[list[str], list[list[str]], str]:
    """行＝状態 / 列＝イベントの表。適用できない場合は理由を返す。

    画面の遷移表タブと同じ内容。イベント列は ID ではなくラベルで並べる
    （`link:P002` のままでは何を押したのか読めないため）。
    """
    from graph.state_table import build_state_transition_report

    data = build_state_transition_report(report.get("screens") or []) or {}
    if not data.get("applicable"):
        return [], [], str(data.get("reason") or "状態遷移テストを適用できません。")

    events = data.get("events") or []
    headers = [str(e.get("label") or e.get("event_id") or "") for e in events]
    order = [str(e.get("event_id") or "") for e in events]

    grid: list[list[str]] = []
    for row in data.get("matrix") or []:
        cells = {str(c.get("event_id")): c for c in (row.get("cells") or [])}
        line = [f"{row.get('state_id')} {row.get('title') or ''}".strip()]
        for event_id in order:
            cell = cells.get(event_id) or {}
            to = str(cell.get("to") or "－")
            # 無効遷移は「受け付けない」ことが情報なので、空欄にせず印を付ける
            line.append(to if cell.get("valid") else f"{to}（無効）" if to != "－" else "－")
        grid.append(line)
    return headers, grid, ""


def _write_state_table(ws: openpyxl.worksheet.worksheet.Worksheet, report: dict[str, Any]) -> int:
    events, grid, reason = build_state_table_grid(report)
    if reason:
        # 空のシートを置くと「遷移が無い」のか「作れなかった」のか区別できない。
        ws.append(["遷移表を作成できませんでした"])
        ws.append([reason])
        ws.column_dimensions["A"].width = 60
        return 0
    ws.append(["状態＼イベント", *events])
    for row in grid:
        ws.append(row)
    _style_header(ws, [22, *[20] * len(events)])
    return len(grid)


# ─────────────────────────────── 組み立て ───────────────────────────────


def build_workbook(domain: str, out_dir: Path) -> tuple[openpyxl.Workbook, dict[str, int]]:
    """既存の spec.xlsx に 3 シートを足したブックを組み立てる。

    既存ブックが無い場合も 3 シートだけのブックとして成立させる
    （画面では見られるのにエクスポートだけ失敗する、という状態を避ける）。
    """
    report = _load_report(domain, out_dir)
    existing = out_dir / domain / XLSX_FILE_NAME

    if existing.is_file():
        wb = openpyxl.load_workbook(existing)
        # 作り直しなので、前回足した分は消してから積み直す（重複防止）。
        for name in ADDED_SHEETS:
            if name in wb.sheetnames:
                del wb[name]
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

    counts = {
        SHEET_TEST_DESIGN: _write_test_design(wb.create_sheet(SHEET_TEST_DESIGN), report),
        SHEET_TESTCASES: _write_testcases(wb.create_sheet(SHEET_TESTCASES), domain, report),
        SHEET_STATE_TABLE: _write_state_table(wb.create_sheet(SHEET_STATE_TABLE), report),
    }
    return wb, counts


def write_full_spec_xlsx(domain: str, out_dir: Path) -> tuple[Path, dict[str, int]]:
    """ブックを組み立てて `spec.xlsx` へ書き戻し、パスと各シートの行数を返す。"""
    wb, counts = build_workbook(domain, out_dir)
    target = out_dir / domain / XLSX_FILE_NAME
    target.parent.mkdir(parents=True, exist_ok=True)
    wb.save(target)
    return target, counts
