"""文書駆動MBTパスから人が読めるテスト手順を生成する。"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

HEADERS = (
    "ケースID",
    "要件ID",
    "タイトル",
    "ステップ",
    "操作",
    "期待結果（レビュー必須）",
    "画面ID",
    "スクリーンショット",
    "根拠",
)


def build_manual_procedures(model: dict[str, Any], report: dict[str, Any]) -> list[dict[str, Any]]:
    """MBTパスを、実測画面へ追跡可能な手動テストケースへ変換する。"""
    screens = {
        str(screen.get("page_id", "")): screen
        for screen in _dict_items(report.get("screens", []))
        if str(screen.get("page_id", ""))
    }
    procedures: list[dict[str, Any]] = []
    for path in _dict_items(model.get("paths", [])):
        node_ids = [str(item) for item in path.get("node_ids", []) if str(item)]
        if not node_ids:
            continue
        steps = _steps(node_ids, screens)
        titles = [_screen_title(screens.get(node_id), node_id) for node_id in node_ids]
        procedures.append(
            {
                "case_id": str(path.get("path_id", "")),
                "title": " → ".join(titles),
                "requirement_ids": [
                    str(item) for item in path.get("requirement_ids", []) if str(item)
                ],
                "preconditions": ["対象サイトへアクセスできること"],
                "steps": steps,
                "evidence": "measured_path",
                "review_required": True,
            }
        )
    return procedures


def save_manual_procedures(procedures: list[dict[str, Any]], output_dir: Path) -> dict[str, Path]:
    """手動テスト手順をMarkdown / Excelへ保存する。"""
    output_dir.mkdir(parents=True, exist_ok=True)
    markdown_path = output_dir / "manual_test_procedures.md"
    excel_path = output_dir / "manual_test_procedures.xlsx"
    markdown_path.write_text(_render_markdown(procedures), encoding="utf-8")
    _write_excel(procedures, excel_path)
    return {
        "manual_procedures_md": markdown_path,
        "manual_procedures_xlsx": excel_path,
    }


def _render_markdown(procedures: list[dict[str, Any]]) -> str:
    lines = [
        "# 手動テスト手順書",
        "",
        "> 画面とパスは実測に基づきます。期待結果は生成物であり、実行前レビューが必要です。",
        "",
    ]
    for procedure in procedures:
        case_id = str(procedure.get("case_id", ""))
        title = str(procedure.get("title", ""))
        requirement_ids = ", ".join(str(item) for item in procedure.get("requirement_ids", []))
        lines.extend(
            [
                f"## {case_id} {title}",
                "",
                f"- 要件ID: {requirement_ids or '-'}",
                f"- 事前条件: {', '.join(procedure.get('preconditions', []))}",
                "- 根拠: measured_path",
                "",
                "| # | 操作 | 期待結果（レビュー必須） | 画面 |",
                "|---:|---|---|---|",
            ]
        )
        screenshots: list[tuple[str, str]] = []
        for step in _dict_items(procedure.get("steps", [])):
            page_id = str(step.get("page_id", ""))
            screenshot = str(step.get("screenshot_path", ""))
            lines.append(
                "| {} | {} | {} | {} |".format(
                    step.get("step_no", ""),
                    _table_cell(step.get("action", "")),
                    _table_cell(step.get("expected_result", "")),
                    _table_cell(page_id),
                )
            )
            if screenshot:
                screenshots.append((page_id, screenshot))
        lines.append("")
        for page_id, screenshot in screenshots:
            lines.extend([f"![{page_id}]({screenshot})", ""])
    return "\n".join(lines)


def _write_excel(procedures: list[dict[str, Any]], path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "手動テスト手順"
    sheet.append(HEADERS)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDEBF7")
    for procedure in procedures:
        requirement_ids = ", ".join(str(item) for item in procedure.get("requirement_ids", []))
        for step in _dict_items(procedure.get("steps", [])):
            screenshot = _excel_text(step.get("screenshot_path", ""))
            sheet.append(
                (
                    _excel_text(procedure.get("case_id", "")),
                    _excel_text(requirement_ids),
                    _excel_text(procedure.get("title", "")),
                    step.get("step_no", ""),
                    _excel_text(step.get("action", "")),
                    _excel_text(step.get("expected_result", "")),
                    _excel_text(step.get("page_id", "")),
                    screenshot,
                    _excel_text(procedure.get("evidence", "")),
                )
            )
            if screenshot:
                screenshot_cell = sheet.cell(row=sheet.max_row, column=8)
                screenshot_cell.hyperlink = screenshot
                screenshot_cell.style = "Hyperlink"
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column, width in {"A": 16, "B": 20, "C": 28, "D": 10, "E": 48, "F": 48}.items():
        sheet.column_dimensions[column].width = width
    workbook.save(path)


def _table_cell(value: object) -> str:
    return str(value).replace("|", "\\|").replace("\n", "<br>")


def _excel_text(value: object) -> str:
    """文書由来の値をExcel数式として解釈させない。"""
    text = str(value)
    return f"'{text}" if text.startswith(("=", "+", "-", "@")) else text


def _steps(node_ids: list[str], screens: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    first_id = node_ids[0]
    first = screens.get(first_id, {})
    first_title = _screen_title(first, first_id)
    steps = [
        {
            "step_no": 1,
            "action": f"{str(first.get('url', ''))} を開く",
            "expected_result": f"「{first_title}」が表示される（生成結果・レビュー必須）",
            "page_id": first_id,
            "screenshot_path": str(first.get("screenshot_path", "")),
        }
    ]
    for index, (source_id, target_id) in enumerate(zip(node_ids, node_ids[1:], strict=False), 2):
        source_title = _screen_title(screens.get(source_id), source_id)
        target = screens.get(target_id, {})
        target_title = _screen_title(target, target_id)
        steps.append(
            {
                "step_no": index,
                "action": f"「{source_title}」から「{target_title}」へ遷移する",
                "expected_result": f"「{target_title}」が表示される（生成結果・レビュー必須）",
                "page_id": target_id,
                "screenshot_path": str(target.get("screenshot_path", "")),
            }
        )
    return steps


def _screen_title(screen: dict[str, Any] | None, fallback: str) -> str:
    return str((screen or {}).get("title") or fallback)


def _dict_items(value: object) -> list[dict[str, Any]]:
    if not isinstance(value, list):
        return []
    return [item for item in value if isinstance(item, dict)]
