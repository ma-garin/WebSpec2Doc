"""SPEC-4-1: 項目定義書（Excel）＋境界値分析（BVA）テストデータ自動生成のテスト。

対象:
    - src/analyzer/bva.py::derive_boundary_cases / attach_observed_boundary_cases
    - src/main.py::_save_excel_output（項目定義書・境界値データシート）
"""

from __future__ import annotations

import re
from pathlib import Path

import openpyxl
from networkx import DiGraph

from analyzer.bva import attach_observed_boundary_cases, derive_boundary_cases
from analyzer.html_analyzer import analyze_pages
from crawler.page_crawler import (
    FieldData,
    FormData,
    PageData,
    SourceEvidence,
    ValidationObservation,
)


def _field(
    name: str = "field1",
    field_type: str = "text",
    required: bool = False,
    maxlength: int | None = None,
    minlength: int | None = None,
    min_value: str = "",
    max_value: str = "",
    pattern: str = "",
    options: tuple[str, ...] = (),
) -> FieldData:
    return FieldData(
        field_type=field_type,
        name=name,
        placeholder="",
        required=required,
        maxlength=maxlength,
        minlength=minlength,
        min_value=min_value,
        max_value=max_value,
        pattern=pattern,
        options=options,
        evidence=SourceEvidence(selector=f"[name='{name}']"),
    )


class TestMaxLengthCases:
    def test_maxlength_three_cases(self) -> None:
        field = _field(maxlength=50)
        cases = derive_boundary_cases(field)
        max_len_cases = [c for c in cases if c.kind == "max_length"]
        assert len(max_len_cases) == 3
        assert [len(c.value) for c in max_len_cases] == [49, 50, 51]
        assert [c.expected for c in max_len_cases] == ["受理", "受理", "エラー（最大長超過）"]
        assert all(c.source_attribute == "maxlength" for c in max_len_cases)

    def test_large_maxlength_does_not_embed_giant_literal(self) -> None:
        """Excel セル上限（32,767文字）を超えるような値はリテラルを埋め込まない。"""
        field = _field(maxlength=50000)
        cases = [c for c in derive_boundary_cases(field) if c.kind == "max_length"]
        assert len(cases) == 3
        for case in cases:
            assert len(case.value) < 32767
        assert "50001" in cases[2].value

    def test_evidence_propagated(self) -> None:
        field = _field(maxlength=10)
        cases = derive_boundary_cases(field)
        assert all(c.evidence is field.evidence for c in cases)


class TestMinLengthCases:
    def test_minlength_two_cases(self) -> None:
        field = _field(minlength=5)
        cases = [c for c in derive_boundary_cases(field) if c.kind == "min_length"]
        assert len(cases) == 2
        assert [len(c.value) for c in cases] == [4, 5]
        assert [c.expected for c in cases] == ["エラー（最小長未満）", "受理"]

    def test_minlength_one_generates_no_case(self) -> None:
        """minlength=1 では空値との比較になり validity.tooShort が発火しないため、
        根拠のある「最小長未満」ケースを生成できない。"""
        field = _field(minlength=1)
        cases = [c for c in derive_boundary_cases(field) if c.kind == "min_length"]
        assert cases == []

    def test_minlength_zero_generates_no_case(self) -> None:
        field = _field(minlength=0)
        cases = [c for c in derive_boundary_cases(field) if c.kind == "min_length"]
        assert cases == []


class TestRangeCases:
    def test_range_boundaries(self) -> None:
        field = _field(field_type="number", min_value="100", max_value="1000000")
        cases = derive_boundary_cases(field)
        range_cases = [c for c in cases if c.kind in ("range_min", "range_max")]
        assert len(range_cases) == 4
        values = [c.value for c in range_cases]
        assert values == ["99", "100", "1000000", "1000001"]
        assert [c.expected for c in range_cases] == [
            "エラー（範囲未満）",
            "受理",
            "受理",
            "エラー（範囲超過）",
        ]


class TestPatternCases:
    def test_pattern_known_generates(self) -> None:
        field = _field(pattern="[0-9]{3,4}")
        cases = [c for c in derive_boundary_cases(field) if c.kind.startswith("pattern")]
        assert len(cases) == 2
        valid = next(c for c in cases if c.kind == "pattern_valid")
        invalid = next(c for c in cases if c.kind == "pattern_invalid")
        assert valid.generated is True
        assert re.fullmatch(field.pattern, valid.value)
        assert re.fullmatch(field.pattern, invalid.value) is None

    def test_pattern_unknown_not_fabricated(self) -> None:
        field = _field(pattern=r"(?=.*[A-Z])(?=.*\d).{8,}")
        cases = [c for c in derive_boundary_cases(field) if c.kind.startswith("pattern")]
        assert len(cases) == 1
        assert cases[0].generated is False
        assert cases[0].value == ""
        assert "手動作成" in cases[0].expected

    def test_card_number_pattern(self) -> None:
        """デモサイト checkout.html のカード番号パターンで適合/不適合例が生成される。"""
        field = _field(pattern="[0-9 ]{13,19}")
        cases = [c for c in derive_boundary_cases(field) if c.kind.startswith("pattern")]
        valid = next(c for c in cases if c.kind == "pattern_valid")
        invalid = next(c for c in cases if c.kind == "pattern_invalid")
        assert re.fullmatch(field.pattern, valid.value)
        assert re.fullmatch(field.pattern, invalid.value) is None

    def test_month_year_pattern(self) -> None:
        field = _field(pattern=r"(0[1-9]|1[0-2])/[0-9]{2}")
        cases = [c for c in derive_boundary_cases(field) if c.kind.startswith("pattern")]
        valid = next(c for c in cases if c.kind == "pattern_valid")
        invalid = next(c for c in cases if c.kind == "pattern_invalid")
        assert re.fullmatch(field.pattern, valid.value)
        assert re.fullmatch(field.pattern, invalid.value) is None


class TestRequiredCases:
    def test_required_empty_case(self) -> None:
        field = _field(required=True)
        cases = [c for c in derive_boundary_cases(field) if c.kind == "required_empty"]
        assert len(cases) == 1
        assert cases[0].value == ""
        assert cases[0].expected == "エラー（必須）"

    def test_required_uses_observed_message(self) -> None:
        field = _field(name="email", required=True)
        cases = derive_boundary_cases(field)
        observation = ValidationObservation(
            field_name="email", message="このフィールドを入力してください"
        )
        updated = attach_observed_boundary_cases(cases, field, [observation])
        required_case = next(c for c in updated if c.kind == "required_empty")
        assert required_case.expected == "このフィールドを入力してください"
        assert required_case.confidence == 1.0

    def test_not_required_no_case(self) -> None:
        field = _field(required=False)
        cases = [c for c in derive_boundary_cases(field) if c.kind == "required_empty"]
        assert cases == []


class TestOptionCases:
    def test_option_first_and_last(self) -> None:
        field = _field(options=("A", "B", "C"), required=True)
        cases = [c for c in derive_boundary_cases(field) if c.kind == "option"]
        assert [c.value for c in cases] == ["A", "C"]

    def test_option_includes_empty_when_not_required(self) -> None:
        field = _field(options=("A", "B"), required=False)
        cases = [c for c in derive_boundary_cases(field) if c.kind == "option"]
        assert [c.value for c in cases] == ["A", "B", ""]


class TestEdgeCases:
    def test_maxlength_zero_skipped(self) -> None:
        field = _field(maxlength=0)
        assert [c for c in derive_boundary_cases(field) if c.kind == "max_length"] == []

    def test_unparseable_range_skipped(self) -> None:
        field = _field(field_type="number", min_value="abc", max_value="100")
        assert [c for c in derive_boundary_cases(field) if c.kind.startswith("range")] == []


def _page(url: str, title: str, fields: tuple[FieldData, ...] = ()) -> PageData:
    forms = (FormData(action="/submit", method="post", fields=fields),) if fields else ()
    return PageData(
        url=url, title=title, headings=(title,), links=(), forms=forms, screenshot_path=None
    )


class TestExcelSheets:
    def test_excel_sheets_added(self, tmp_path: Path) -> None:
        from main import _save_excel_output

        page = _page(
            "https://example.com/checkout",
            "決済",
            (
                _field(
                    name="amount",
                    field_type="number",
                    min_value="100",
                    max_value="1000000",
                    required=True,
                ),
            ),
        )
        analyzed = analyze_pages([page])
        _save_excel_output(tmp_path, analyzed, [], official_names=None)

        wb = openpyxl.load_workbook(tmp_path / "spec.xlsx")
        assert wb.sheetnames == ["Screens", "Forms", "項目定義書", "境界値データ"]
        field_def_ws = wb["項目定義書"]
        assert field_def_ws.cell(row=1, column=1).value == "画面名"
        bva_ws = wb["境界値データ"]
        assert bva_ws.cell(row=1, column=1).value == "画面ID"

    def test_official_name_injected(self, tmp_path: Path) -> None:
        from main import _save_excel_output

        page = _page(
            "https://example.com/checkout", "決済", (_field(name="amount", required=True),)
        )
        analyzed = analyze_pages([page])
        page_id = analyzed[0].page_id
        _save_excel_output(tmp_path, analyzed, [], official_names={page_id: "与信申込入力"})

        wb = openpyxl.load_workbook(tmp_path / "spec.xlsx")
        ws = wb["項目定義書"]
        assert ws.cell(row=2, column=1).value == "与信申込入力"

    def test_report_json_unchanged(self) -> None:
        """BVA/項目定義書はExcel専用の出力であり、report.jsonにキーを追加しない。"""
        import json

        from generator.json_reporter import generate_json_report

        page = _page(
            "https://example.com/checkout", "決済", (_field(name="amount", required=True),)
        )
        analyzed = analyze_pages([page])
        graph = DiGraph()
        graph.add_node(analyzed[0].page_id)
        data = json.loads(generate_json_report(analyzed, graph, "https://example.com/"))
        field_dict = data["screens"][0]["forms"][0]["fields"][0]
        assert "boundary_cases" not in field_dict
        assert "field_definition" not in data["screens"][0]
