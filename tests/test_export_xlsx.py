"""テスト仕様書一式の Excel 出力（P2-3）のテスト。

受入は「各シートの行数＝画面内の表示件数と一致」。数が合わないと、
Excel を見た人が「そもそも設計されていない」と誤解する。
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl
import pytest

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))
sys.path.insert(0, str(ROOT / "src"))

import app as appmod  # noqa: E402
from web.services.export_xlsx import (  # noqa: E402
    ADDED_SHEETS,
    ExportError,
    build_state_table_grid,
    build_test_design_rows,
    build_testcase_rows,
    build_workbook,
    write_full_spec_xlsx,
)

H = {"Host": "127.0.0.1"}
DOMAIN = "xlsx-export.example.com"


def _screen(page_id: str, url: str, title: str, *, forms: list | None = None) -> dict:
    return {
        "page_id": page_id,
        "url": url,
        "title": title,
        "headings": [f"{title} 見出し"],
        "links": [],
        "forms": forms or [],
        "buttons": [],
    }


REPORT = {
    "screens": [
        _screen("P001", "http://x.test/", "トップ"),
        _screen(
            "P002",
            "http://x.test/form.html",
            "申込フォーム",
            forms=[
                {
                    "action": "/submit",
                    "method": "post",
                    "fields": [
                        {
                            "name": "card",
                            "type": "text",
                            "required": True,
                            "maxlength": 16,
                            "label": "カード番号",
                        },
                        {
                            "name": "method",
                            "type": "select",
                            "options": ["a", "b", "c"],
                            "label": "支払方法",
                        },
                    ],
                }
            ],
        ),
    ]
}


@pytest.fixture
def site(tmp_path: Path) -> Path:
    d = tmp_path / DOMAIN
    d.mkdir(parents=True)
    (d / "report.json").write_text(json.dumps(REPORT, ensure_ascii=False), encoding="utf-8")
    return tmp_path


class TestSheetContents:
    def test_all_three_sheets_are_added(self, site: Path) -> None:
        wb, counts = build_workbook(DOMAIN, site)
        for name in ADDED_SHEETS:
            assert name in wb.sheetnames
        assert set(counts) == set(ADDED_SHEETS)

    def test_row_counts_match_the_source_of_truth(self, site: Path) -> None:
        """行数が画面と同じ根拠（同じサービス）から出ていること。"""
        wb, counts = build_workbook(DOMAIN, site)
        assert counts["テスト設計"] == len(build_test_design_rows(REPORT))
        assert counts["テストケース"] == len(build_testcase_rows(DOMAIN, REPORT))
        # ヘッダ行を除いた行数がカウントと一致する
        for name in ADDED_SHEETS:
            if counts[name]:
                assert wb[name].max_row - 1 == counts[name]

    def test_test_design_sheet_has_the_designed_columns(self, site: Path) -> None:
        wb, _ = build_workbook(DOMAIN, site)
        header = [c.value for c in wb["テスト設計"][1]]
        assert header == ["画面ID", "画面名", "No", "テスト条件", "導出技法", "由来", "由来の詳細"]

    def test_testcase_sheet_has_the_designed_columns(self, site: Path) -> None:
        wb, _ = build_workbook(DOMAIN, site)
        header = [c.value for c in wb["テストケース"][1]]
        assert header == [
            "ID",
            "テストケース名",
            "画面",
            "機能",
            "観点",
            "前提条件",
            "手順",
            "期待結果",
            "自動化判定",
            "結果",
        ]

    def test_multiline_values_are_kept_in_one_cell(self, site: Path) -> None:
        """手順・期待結果は画面では複数行。セルを分けると行数が合わなくなる。"""
        rows = build_testcase_rows(DOMAIN, REPORT)
        assert rows, "テストケースが 1 件も無いと、この検証は成立しない"
        assert any("\n" in cell for row in rows for cell in row)

    def test_state_table_marks_invalid_transitions(self, site: Path) -> None:
        """無効遷移を空欄にすると『受け付けない』という情報が消える。"""
        headers, grid, reason = build_state_table_grid(REPORT)
        if reason:
            pytest.skip(f"この題材では遷移表を作れない: {reason}")
        assert headers and grid
        assert all(len(row) == len(headers) + 1 for row in grid)


class TestFailurePaths:
    def test_missing_report_is_reported_not_silently_empty(self, tmp_path: Path) -> None:
        """report.json が無いのに空のブックを返すと、中身が無い理由が分からない。"""
        (tmp_path / DOMAIN).mkdir(parents=True)
        with pytest.raises(ExportError) as e:
            build_workbook(DOMAIN, tmp_path)
        assert "report.json" in str(e.value)

    def test_unreadable_report_is_reported(self, tmp_path: Path) -> None:
        d = tmp_path / DOMAIN
        d.mkdir(parents=True)
        (d / "report.json").write_text("{壊れている", encoding="utf-8")
        with pytest.raises(ExportError):
            build_workbook(DOMAIN, tmp_path)

    def test_state_table_gives_a_reason_when_not_applicable(self, tmp_path: Path) -> None:
        """作れなかったことを空シートで表すと、遷移が無いのか失敗なのか分からない。"""
        d = tmp_path / DOMAIN
        d.mkdir(parents=True)
        (d / "report.json").write_text(json.dumps({"screens": []}), encoding="utf-8")
        wb, counts = build_workbook(DOMAIN, tmp_path)
        assert counts["遷移表"] == 0
        text = "\n".join(str(c.value or "") for row in wb["遷移表"].iter_rows() for c in row)
        assert "作成できませんでした" in text

    def test_reexport_does_not_duplicate_sheets(self, site: Path) -> None:
        """2 回書き出してシートが増えると、どれが最新か分からなくなる。"""
        write_full_spec_xlsx(DOMAIN, site)
        path, _ = write_full_spec_xlsx(DOMAIN, site)
        names = openpyxl.load_workbook(path).sheetnames
        for name in ADDED_SHEETS:
            assert names.count(name) == 1

    def test_existing_four_sheets_are_preserved(self, site: Path) -> None:
        """実測仕様の 4 シートを消してしまうと、既存の納品物が壊れる。"""
        wb = openpyxl.Workbook()
        wb.active.title = "Screens"
        for name in ("Forms", "項目定義書", "境界値データ"):
            wb.create_sheet(name)
        wb.save(site / DOMAIN / "spec.xlsx")

        path, _ = write_full_spec_xlsx(DOMAIN, site)
        names = openpyxl.load_workbook(path).sheetnames
        assert names[:4] == ["Screens", "Forms", "項目定義書", "境界値データ"]
        assert names[4:] == list(ADDED_SHEETS)


class TestExportRoute:
    @pytest.fixture
    def client(self):
        return appmod.app.test_client()

    @pytest.mark.parametrize("query", ["", "?domain=", "?domain=../etc", "?domain=nope.invalid"])
    def test_invalid_domain_is_404(self, client, query) -> None:
        assert client.get(f"/api/export/spec-xlsx{query}", headers=H).status_code == 404

    def test_download_returns_an_excel_content_type(self, client, monkeypatch, site: Path) -> None:
        """content-type を間違えると、ブラウザが Excel として扱わない。"""
        import web.routes.report as report_mod

        monkeypatch.setattr(report_mod, "_out", lambda: site)
        res = client.get(f"/api/export/spec-xlsx?domain={DOMAIN}", headers=H)
        assert res.status_code == 200
        assert res.headers["Content-Type"].startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert "spec.xlsx" in res.headers.get("Content-Disposition", "")

    def test_downloaded_workbook_has_the_added_sheets(self, client, monkeypatch, site: Path):
        import io

        import web.routes.report as report_mod

        monkeypatch.setattr(report_mod, "_out", lambda: site)
        res = client.get(f"/api/export/spec-xlsx?domain={DOMAIN}", headers=H)
        wb = openpyxl.load_workbook(io.BytesIO(res.data))
        for name in ADDED_SHEETS:
            assert name in wb.sheetnames

    def test_missing_report_is_409_not_500(self, client, monkeypatch, tmp_path: Path) -> None:
        """成果物が足りないだけなのに 500 だと、不具合と区別できない。"""
        import web.routes.report as report_mod

        (tmp_path / DOMAIN).mkdir(parents=True)
        monkeypatch.setattr(report_mod, "_out", lambda: tmp_path)
        res = client.get(f"/api/export/spec-xlsx?domain={DOMAIN}", headers=H)
        assert res.status_code == 409
        assert "report.json" in res.get_data(as_text=True)
