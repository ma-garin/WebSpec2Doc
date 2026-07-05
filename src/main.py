from __future__ import annotations

import argparse
import json
import logging
import os
import signal
import sys
import threading
from collections.abc import Callable, Sequence
from datetime import UTC, datetime
from pathlib import Path
from typing import TYPE_CHECKING
from urllib.parse import urlparse

import networkx as nx
import openpyxl
from dotenv import load_dotenv

from analyzer.form_analyzer import summarize_forms
from analyzer.html_analyzer import AnalyzedPage, analyze_pages
from crawler.auth import DEFAULT_AUTH_FILE, capture_auth_state
from crawler.page_crawler import (
    DEFAULT_DEPTH,
    DEFAULT_MAX_PAGES,
    PageData,
    crawl_site,
    crawl_urls,
    discover_pages,
    evidence_to_dict,
)
from crawler.session_guard import SessionExpiredError
from diff.differ import compute_diff
from diff.snapshot import latest_snapshot, load_snapshot, save_partial_snapshot, save_snapshot
from generator.diff_reporter import generate_diff_report
from generator.markdown_generator import (
    generate_forms_markdown,
    generate_screens_markdown,
)
from generator.mermaid_generator import generate_mermaid
from graph.transition_graph import build_graph

if TYPE_CHECKING:
    from ux.axe_runner import AxeViolation

DEFAULT_OUTPUT_DIR = Path("output")
DEFAULT_FORMATS = "md"
FORMAT_SEPARATOR = ","
LOGGER_FORMAT = "%(levelname)s:%(name)s:%(message)s"
SUPPORTED_FORMATS = frozenset({"md", "html", "excel", "pdf", "json"})
XLSX_FILE_NAME = "spec.xlsx"
PDF_FILE_NAME = "report.pdf"
JSON_REPORT_FILE_NAME = "report.json"
DIFF_REPORT_FILE_NAME = "diff_report.html"
FIRST_SNAPSHOT_MESSAGE = "初回スナップショットを保存しました。次回実行から差分を検出できます"

logger = logging.getLogger(__name__)
_STOP_REQUESTED = threading.Event()
_EVENT_WRITE_LOCK = threading.Lock()


def main() -> None:
    logging.basicConfig(level=logging.INFO, format=LOGGER_FORMAT)
    load_dotenv()
    signal.signal(signal.SIGTERM, _request_stop)
    signal.signal(signal.SIGINT, _request_stop)
    run(parse_args())


def _request_stop(_signum: int, _frame: object) -> None:
    _STOP_REQUESTED.set()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="WebSpec2Doc web crawler")
    parser.add_argument("--url", help="クロール対象URL")
    parser.add_argument(
        "--urls",
        help="リンク追跡せず指定URLのみクロール（カンマ区切り）。--url の代わりに使用",
    )
    parser.add_argument(
        "--discover",
        action="store_true",
        help="リンク追跡で到達するページ一覧(URL+タイトル)をJSONでstdoutに出力して終了",
    )
    parser.add_argument(
        "--stream",
        action="store_true",
        help="--discover と組み合わせ、発見ページを NDJSON でリアルタイム出力する（GUI用）",
    )
    parser.add_argument("--login", help="ログインセッション保存モード: ログインページURL（CLI用）")
    parser.add_argument(
        "--login-signal",
        type=Path,
        help="GUI 手渡しログイン用: このファイルが出現したらセッションを保存する（廃止予定）",
    )
    parser.add_argument(
        "--login-simple", action="store_true", help="ID/PASSWORDをstdinのJSONで受取り自動ログイン"
    )
    parser.add_argument("--login-simple-url", help="--login-simple: ログインページURL")
    parser.add_argument(
        "--login-scrape", help="ログインURLのフォームフィールドを取得してJSONで出力"
    )
    parser.add_argument(
        "--login-submit",
        action="store_true",
        help="ログインフォーム自動送信（フィールド値はstdinからJSON受取）",
    )
    parser.add_argument("--login-current-url", help="--login-submit: フォーム送信先URL（MFA対応）")
    parser.add_argument("--login-temp-session", type=Path, help="MFA中間セッション保存パス")
    parser.add_argument(
        "--login-record",
        action="store_true",
        help=(
            "認証フローレコーダー起動: 見えるブラウザでログインし、"
            "シグナルファイル出現時にセッションを保存する（--login と --login-signal の統合入口）"
        ),
    )
    parser.add_argument("--login-record-url", help="--login-record: ログインページURL")
    parser.add_argument(
        "--login-status",
        type=Path,
        help="--login-record: 進行状態(JSON)の出力先。Web UI が1秒間隔でポーリングする",
    )
    parser.add_argument("--auth", type=Path, help="保存済みセッション(auth.json)を使ってクロール")
    parser.add_argument("--depth", type=int, default=DEFAULT_DEPTH, help="クロール深度")
    parser.add_argument(
        "--max-pages",
        type=int,
        default=DEFAULT_MAX_PAGES,
        help="最大クロールページ数",
    )
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT_DIR, help="出力先")
    parser.add_argument("--llm", action="store_true", help="LLM解析を有効化")
    parser.add_argument(
        "--format", default=DEFAULT_FORMATS, help="出力形式: md,html,excel,pdf,json"
    )
    parser.add_argument("--compare", action="store_true", help="前回スナップショットとの差分を出力")
    parser.add_argument(
        "--parallelism",
        type=int,
        default=1,
        help="明示URLクロールの並列数（GUI既定: 2、最大: 4）",
    )
    parser.add_argument(
        "--fail-on-drift",
        action="store_true",
        help="差分検知時に exit code 1 で終了（CI/CDパイプライン用）",
    )
    parser.add_argument(
        "--reference-doc",
        type=Path,
        action="append",
        help=(
            "参考文書（既存の画面一覧・項目定義書など）。実測結果と突合し"
            "ギャップレポートと正式画面名の注入を行う。複数指定可。"
            "対応形式: xlsx/docx/pptx/pdf/md/txt/yaml/json"
        ),
    )
    parser.add_argument(
        "--record-session",
        action="store_true",
        help=(
            "探索セッション記録モード: --url のページを記録用ブラウザで開き、"
            "閉じられるまで操作（クリック・入力・遷移・画面状態）を記録する"
        ),
    )
    parser.add_argument(
        "--record-duration",
        type=float,
        default=None,
        help="--record-session の最大記録時間（秒）。未指定はブラウザを閉じるまで",
    )
    parser.add_argument(
        "--exploration-coverage",
        action="store_true",
        help=(
            "探索カバレッジ集計モード: クロール済み report.json と記録済み"
            "セッションから exploration_heatmap.html 等を生成する"
        ),
    )
    parser.add_argument(
        "--reverse-assets",
        action="store_true",
        help=(
            "リバース生成モード: クロール済み report.json と記録済みセッションから"
            "テストケース・記録フロー（recorded_assets.json / recorded_candidates.json）を"
            "逆生成する"
        ),
    )
    parser.add_argument(
        "--export-findings",
        action="store_true",
        help=(
            "気づき票エクスポートモード: 記録済みセッションの気づきマーク（finding イベント）を"
            "再現手順付きバグ票として findings.json / findings.csv に出力する"
        ),
    )
    parser.add_argument(
        "--doc-llm",
        action="store_true",
        help=(
            "--reference-doc の自由文形式（pdf/pptx/txt/docx 本文）から LLM で"
            "画面・項目・業務ルールを追加抽出する（既定 OFF。OPENAI_API_KEY 必須）"
        ),
    )
    parser.add_argument(
        "--test-plan",
        action="store_true",
        help=(
            "テスト計画ドラフト生成モード: クロール済み report.json から"
            "画面数×優先度の工数見積・スコープ表（test_plan.md / test_plan.xlsx）を生成する"
        ),
    )
    parser.add_argument(
        "--compare-old-urls",
        help=("現新比較モード: 現行側 URL（カンマ区切り）。--compare-new-urls と併用する"),
    )
    parser.add_argument(
        "--compare-new-urls",
        help=("現新比較モード: 新側 URL（カンマ区切り）。--compare-old-urls と併用する"),
    )
    parser.add_argument(
        "--compare-auth-old",
        type=Path,
        help="現新比較: 現行側の保存済みセッション(auth.json)",
    )
    parser.add_argument(
        "--compare-auth-new",
        type=Path,
        help="現新比較: 新側の保存済みセッション(auth.json)",
    )
    parser.add_argument(
        "--compare-mask-selector",
        action="append",
        help=(
            "現新比較: 動的領域として画像差分から除外する CSS セレクタ"
            "（広告枠など既知の動的領域）。複数指定可"
        ),
    )
    parser.add_argument(
        "--ux-review",
        action="store_true",
        help=(
            "UX自動エキスパートレビュー: axe-core によるWCAG違反検査とニールセン10原則の"
            "ヒューリスティック評価を行い ux_review.json / report.html「UX所見」タブを生成する"
            "（OPENAI_API_KEY 未設定時は rules ベースの評価で完走する）"
        ),
    )
    parser.add_argument(
        "--refresh-doc",
        action="store_true",
        help=(
            "文書の再生: 参考文書の構造を骨格として維持したまま実測値で更新した"
            "新版仕様書（refreshed_spec.md）と変更ログ（refresh_log.json）を生成する"
            "（既定 OFF。--reference-doc と併用必須。単独指定は警告して無視）"
        ),
    )
    return parser.parse_args()


def run(args: argparse.Namespace) -> None:
    auth_path = getattr(args, "auth", None)
    if getattr(args, "login_simple", False):
        _submit_login_simple(args)
        return
    if login_scrape_url := getattr(args, "login_scrape", None):
        _scrape_login(str(login_scrape_url))
        return
    if getattr(args, "login_submit", False):
        _submit_login(args)
        return
    if getattr(args, "login_record", False):
        _record_login(args)
        return
    login_url = getattr(args, "login", None)
    if login_url:
        _capture_login(str(login_url), auth_path, getattr(args, "login_signal", None))
        return
    if bool(getattr(args, "discover", False)):
        _discover(args, auth_path)
        return
    if bool(getattr(args, "record_session", False)):
        _record_session(args)
        return
    if bool(getattr(args, "exploration_coverage", False)):
        _exploration_coverage(args)
        return
    if bool(getattr(args, "reverse_assets", False)):
        _reverse_assets(args)
        return
    if bool(getattr(args, "export_findings", False)):
        _export_findings(args)
        return
    if bool(getattr(args, "test_plan", False)):
        _generate_test_plan(args)
        return
    if getattr(args, "compare_old_urls", None) and getattr(args, "compare_new_urls", None):
        _run_old_new_comparison(args)
        return
    _run_crawl(args, auth_path)


def _run_old_new_comparison(args: argparse.Namespace) -> None:
    """現新比較モード: 2 ターゲットクロール→対応付け→三層比較→4 分類レポート出力。"""
    from diff.comparison import ComparisonError, run_old_new_comparison
    from generator.comparison_reporter import save_comparison_outputs

    old_urls = _parse_url_list(getattr(args, "compare_old_urls", None))
    new_urls = _parse_url_list(getattr(args, "compare_new_urls", None))
    if not old_urls or not new_urls:
        logger.error("--compare-old-urls と --compare-new-urls の両方を指定してください")
        return

    output_dir = (
        Path(args.output) / f"compare_{_domain_name(old_urls[0])}_vs_{_domain_name(new_urls[0])}"
    )
    auth_old = getattr(args, "compare_auth_old", None)
    auth_new = getattr(args, "compare_auth_new", None)
    mask_selectors = tuple(getattr(args, "compare_mask_selector", None) or ())
    _STOP_REQUESTED.clear()

    try:
        result = run_old_new_comparison(
            old_urls,
            new_urls,
            output_dir,
            auth_old=Path(auth_old) if auth_old else None,
            auth_new=Path(auth_new) if auth_new else None,
            mask_selectors=mask_selectors,
            stop_requested=_STOP_REQUESTED.is_set,
        )
    except ComparisonError as exc:
        logger.error("%s", exc)
        sys.exit(1)

    json_path, html_path = save_comparison_outputs(result, output_dir)
    logger.info(
        "現新比較が完了しました: 対応画面 %d 組・指摘 %d 件・新規追加 %d 件・削除 %d 件 — %s / %s",
        len(result.pairs),
        len(result.findings),
        len(result.added_page_ids),
        len(result.removed_page_ids),
        json_path,
        html_path,
    )
    if not result.pairs:
        logger.warning("対応画面が見つかりません")


def _record_session(args: argparse.Namespace) -> None:
    """探索セッション記録モード（キャプチャ Phase 1）。"""
    from capture.session_recorder import record_exploration_session

    url = str(args.url or "")
    if not url:
        logger.error("--record-session には --url が必要です")
        return
    output_dir = Path(args.output) / _domain_name(url)
    headless = os.environ.get("WEBSPEC2DOC_RECORD_HEADLESS", "") == "1"
    logger.info("記録を開始します。ブラウザを閉じるとセッションを保存します: %s", url)
    session_path = record_exploration_session(
        url,
        output_dir,
        duration_sec=getattr(args, "record_duration", None),
        headless=headless,
    )
    logger.info("探索セッションを保存しました: %s", session_path)


def _load_report_json(report_path: Path) -> dict | None:
    """クロール済み report.json を寛容に読み込む。

    不在・破損（JSONDecodeError）・読み込み失敗（OSError）はエラーログを出して
    None を返す。中断で切れた report.json でもスタックトレースを吐かず、各モードが
    一様に「インベントリなし」として安全に中断できるようにする（全モード共通）。
    """
    if not report_path.exists():
        logger.error(
            "クロール済みインベントリがありません: %s（先に --format json でクロールしてください）",
            report_path,
        )
        return None
    try:
        return json.loads(report_path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        logger.error("report.json を読み込めません: %s (%s)", report_path, exc)
        return None


def _exploration_coverage(args: argparse.Namespace) -> None:
    """探索カバレッジ集計モード（ヒートマップ Phase 1）。"""
    from capture.coverage import (
        compute_exploration_coverage,
        load_session_events,
        save_exploration_coverage,
    )

    url = str(args.url or "")
    if not url:
        logger.error("--exploration-coverage には --url が必要です")
        return
    output_dir = Path(args.output) / _domain_name(url)
    report_path = output_dir / JSON_REPORT_FILE_NAME
    report = _load_report_json(report_path)
    if report is None:
        return
    events = load_session_events(output_dir)
    if not events:
        logger.error("探索セッションがありません（先に --record-session で操作を記録してください）")
        return
    business_flows = report.get("meta", {}).get("business_flows")
    coverage = compute_exploration_coverage(report, events, business_flows=business_flows)
    save_exploration_coverage(coverage, output_dir)
    summary = coverage["summary"]
    logger.info(
        "探索カバレッジ: %d/%d 画面（%.0f%%）・状態 %d/%d — exploration_heatmap.html を出力しました",
        summary["explored_screens"],
        summary["total_screens"],
        summary["coverage_ratio"] * 100,
        summary["touched_states"],
        summary["total_states"],
    )

    from capture.burndown import compute_exploration_burndown, save_exploration_burndown
    from capture.session_recorder import SESSIONS_DIR_NAME

    burndown = compute_exploration_burndown(report, events, output_dir / SESSIONS_DIR_NAME)
    save_exploration_burndown(burndown, output_dir)
    logger.info(
        "進捗バーンダウン: %d セッション分の系列を出力しました — exploration_burndown.html",
        burndown["summary"]["session_count"],
    )


def _reverse_assets(args: argparse.Namespace) -> None:
    """リバース生成モード（キャプチャ Phase 2）。"""
    from capture.coverage import load_session_events
    from capture.reverse_generator import generate_recorded_assets, save_recorded_assets

    url = str(args.url or "")
    if not url:
        logger.error("--reverse-assets には --url が必要です")
        return
    output_dir = Path(args.output) / _domain_name(url)
    report_path = output_dir / JSON_REPORT_FILE_NAME
    report = _load_report_json(report_path)
    if report is None:
        return
    events = load_session_events(output_dir)
    if not events:
        logger.error("探索セッションがありません（先に --record-session で操作を記録してください）")
        return
    assets = generate_recorded_assets(report, events)
    save_recorded_assets(assets, output_dir)
    logger.info(
        "リバース生成が完了しました: テストケース %d 件・記録フロー %d 件"
        "（recorded_assets.json / recorded_candidates.json）",
        len(assets["test_cases"]),
        len(assets["flows"]),
    )


def _export_findings(args: argparse.Namespace) -> None:
    """気づき票エクスポートモード（キャプチャ Phase 3: 気づき→バグ票自動起票）。"""
    from capture.coverage import load_session_events
    from capture.finding_reporter import build_finding_tickets, save_findings

    url = str(args.url or "")
    if not url:
        logger.error("--export-findings には --url が必要です")
        return
    output_dir = Path(args.output) / _domain_name(url)
    events = load_session_events(output_dir)
    if not events:
        logger.error("探索セッションがありません（先に --record-session で操作を記録してください）")
        return
    tickets = build_finding_tickets(events)
    save_findings(tickets, output_dir)
    logger.info(
        "気づき票エクスポートが完了しました: %d 件（findings.json / findings.csv）",
        len(tickets),
    )


def _generate_test_plan(args: argparse.Namespace) -> None:
    """テスト計画ドラフト生成モード（計画 Phase 1）。"""
    from generator.test_plan_generator import (
        compute_test_plan,
        load_plan_coefficients,
        save_test_plan,
    )

    url = str(args.url or "")
    if not url:
        logger.error("--test-plan には --url が必要です")
        return
    output_dir = Path(args.output) / _domain_name(url)
    report_path = output_dir / JSON_REPORT_FILE_NAME
    report = _load_report_json(report_path)
    if report is None:
        return
    coefficients = load_plan_coefficients()
    plan = compute_test_plan(report, coefficients)
    save_test_plan(plan, output_dir)
    logger.info(
        "テスト計画ドラフトを生成しました: 画面 %d 件・総見積 %.1f 時間"
        "（test_plan.md / test_plan.xlsx）",
        len(plan.rows),
        plan.total_hours,
    )


def _run_crawl(args: argparse.Namespace, auth_path: Path | None) -> None:
    """メインクロール・分析・出力フローを実行する。"""
    url_list = _parse_url_list(getattr(args, "urls", None))
    primary_url = url_list[0] if url_list else (args.url or "")
    if not primary_url:
        logger.error("--url / --urls / --login のいずれかを指定してください")
        return
    if args.llm:
        logger.warning("--llm は未実装のため無視します")

    ux_review_enabled = bool(getattr(args, "ux_review", False))
    if ux_review_enabled:
        from ux.axe_runner import AxeAssetError, verify_axe_asset

        try:
            verify_axe_asset()
        except AxeAssetError as exc:
            logger.error("%s", exc)
            sys.stdout.write("UX_REVIEW_ASSET_ERROR\n")
            sys.stdout.flush()
            return

    formats = _parse_formats(str(args.format))
    output_dir = Path(args.output) / _domain_name(primary_url)
    prior_snapshot = latest_snapshot(output_dir)
    crawled_at = datetime.now(UTC).strftime("%Y-%m-%d %H:%M UTC")
    auth_state = Path(auth_path) if auth_path else None
    _STOP_REQUESTED.clear()
    checkpoint_pages: list[PageData] = []
    ux_axe_results: dict[str, tuple[AxeViolation, ...]] = {}
    ux_axe_lock = threading.Lock()

    def on_ux_result(url: str, violations: tuple[AxeViolation, ...]) -> None:
        # 並列クロール時は複数ワーカースレッドから呼ばれるためロックする。
        with ux_axe_lock:
            ux_axe_results[url] = violations

    def emit_event(event: dict[str, object]) -> None:
        with _EVENT_WRITE_LOCK:
            sys.stdout.write(f"CRAWL_EVENT:{json.dumps(event, ensure_ascii=False)}\n")
            sys.stdout.flush()

    def save_checkpoint(pages: list[PageData]) -> None:
        nonlocal checkpoint_pages
        checkpoint_pages = list(pages)
        path = save_partial_snapshot(checkpoint_pages, output_dir)
        emit_event(
            {
                "event": "checkpoint_saved",
                "saved_count": len(checkpoint_pages),
                "path": str(path),
            }
        )

    try:
        pages = _do_crawl(
            args,
            url_list,
            primary_url,
            output_dir,
            auth_state,
            on_event=emit_event,
            on_checkpoint=save_checkpoint,
            ux_review=ux_review_enabled,
            on_ux_result=on_ux_result if ux_review_enabled else None,
        )
    except SessionExpiredError as exc:
        if checkpoint_pages:
            partial = save_partial_snapshot(checkpoint_pages, output_dir, finalized=True)
            emit_event(
                {
                    "event": "checkpoint_saved",
                    "saved_count": len(checkpoint_pages),
                    "path": str(partial),
                    "finalized": True,
                }
            )
        logger.error("%s", exc)
        sys.stdout.write("SESSION_EXPIRED\n")
        sys.exit(2)

    analyzed_pages = analyze_pages(pages)
    graph = build_graph(analyzed_pages)
    form_summary = summarize_forms(analyzed_pages)
    transition_coverage, business_flows = _compute_transition_quality(pages)
    impact_report = None
    if bool(getattr(args, "compare", False)) and prior_snapshot is not None:
        impact_report = _compute_impact_report(prior_snapshot, analyzed_pages, output_dir)
    official_names, rule_conditions = _run_doc_fusion(
        analyzed_pages,
        getattr(args, "reference_doc", None),
        output_dir,
        use_llm=bool(getattr(args, "doc_llm", False)),
        refresh_doc=bool(getattr(args, "refresh_doc", False)),
    )
    exploration_coverage = _load_exploration_coverage(output_dir)
    ux_review = None
    if ux_review_enabled:
        ux_review = _build_and_save_ux_review(pages, analyzed_pages, ux_axe_results, output_dir)
    coverage_gaps = _collect_coverage_gaps(output_dir, pages, exploration_coverage)
    save_outputs(
        analyzed_pages,
        graph,
        form_summary,
        output_dir,
        formats,
        crawl_depth=int(args.depth),
        crawl_max_pages=int(args.max_pages),
        crawled_at=crawled_at,
        transition_coverage=transition_coverage,
        business_flows=business_flows,
        impact_report=impact_report,
        official_names=official_names,
        exploration_coverage=exploration_coverage,
        rule_conditions=rule_conditions,
        ux_review=ux_review,
        coverage_gaps=coverage_gaps,
    )
    if _STOP_REQUESTED.is_set():
        partial = save_partial_snapshot(pages, output_dir, finalized=True)
        emit_event(
            {
                "event": "checkpoint_saved",
                "saved_count": len(pages),
                "path": str(partial),
                "finalized": True,
            }
        )
        sys.stdout.write("CRAWL_CANCELLED\n")
        sys.stdout.flush()
        return
    new_snapshot = save_snapshot(pages, output_dir)
    drift_detected = False
    if bool(getattr(args, "compare", False)):
        drift_detected = _save_diff_report(
            prior_snapshot, new_snapshot, pages, output_dir, primary_url
        )
    logger.info("出力が完了しました: %s", output_dir)
    if drift_detected and bool(getattr(args, "fail_on_drift", False)):
        logger.warning(
            "仕様ドリフトを検知しました（--fail-on-drift が有効）。exit code 1 で終了します。"
        )
        sys.exit(1)


def _do_crawl(
    args: argparse.Namespace,
    url_list: list[str],
    primary_url: str,
    output_dir: Path,
    auth_state: Path | None,
    on_event: Callable[[dict[str, object]], None] | None = None,
    on_checkpoint: Callable[[list[PageData]], None] | None = None,
    ux_review: bool = False,
    on_ux_result: Callable[[str, tuple[AxeViolation, ...]], None] | None = None,
) -> list[PageData]:
    """URL リストまたは単一 URL をクロールして PageData リストを返す。"""
    if url_list:
        return crawl_urls(
            url_list,
            output_dir=output_dir,
            auth_state=auth_state,
            parallelism=max(1, min(int(getattr(args, "parallelism", 1)), 4)),
            respect_robots=True,
            on_event=on_event,
            on_checkpoint=on_checkpoint,
            stop_requested=_STOP_REQUESTED.is_set,
            ux_review=ux_review,
            on_ux_result=on_ux_result,
        )
    return crawl_site(
        url=primary_url,
        depth=int(args.depth),
        max_pages=int(args.max_pages),
        output_dir=output_dir,
        auth_state=auth_state,
        on_event=on_event,
        on_checkpoint=on_checkpoint,
        stop_requested=_STOP_REQUESTED.is_set,
        ux_review=ux_review,
        on_ux_result=on_ux_result,
    )


def _parse_url_list(raw_urls: str | None) -> list[str]:
    if not raw_urls:
        return []
    seen: dict[str, None] = {}
    for item in raw_urls.split(FORMAT_SEPARATOR):
        cleaned = item.strip()
        if cleaned:
            seen.setdefault(cleaned, None)
    return list(seen)


def _submit_login_simple(args: argparse.Namespace) -> None:
    """ID/PASSWORDをstdinのJSONで受取り、type属性ベースで自動マッピングしてログインする。"""
    from dataclasses import asdict

    from crawler.auto_login import submit_login_simple

    login_url = getattr(args, "login_simple_url", None) or ""
    if not login_url:
        sys.stdout.write(json.dumps({"success": False, "error": "--login-simple-url が必要です"}))
        return
    auth_path = Path(args.auth) if args.auth else Path(DEFAULT_AUTH_FILE)
    try:
        raw = sys.stdin.read().strip() if not sys.stdin.isatty() else "{}"
        creds: dict[str, str] = json.loads(raw) if raw else {}
    except json.JSONDecodeError:
        sys.stdout.write(json.dumps({"success": False, "error": "認証情報JSONが不正です"}))
        return
    result = submit_login_simple(
        username=creds.get("username", ""),
        password=creds.get("password", ""),
        login_url=login_url,
        auth_path=auth_path,
    )
    sys.stdout.write(
        json.dumps(
            {
                "success": result.success,
                "needs_more_fields": result.needs_more_fields,
                "fields": [asdict(f) for f in result.fields],
                "current_url": result.current_url,
                "error": result.error,
            },
            ensure_ascii=False,
        )
    )


def _scrape_login(url: str) -> None:
    """ログインページのフォームフィールドを取得してJSONをstdoutに出力する。"""
    from dataclasses import asdict

    from crawler.auto_login import scrape_login_fields

    result = scrape_login_fields(url)
    sys.stdout.write(
        json.dumps(
            {
                "ok": result.ok,
                "fields": [asdict(f) for f in result.fields],
                "current_url": result.current_url,
                "error": result.error,
            },
            ensure_ascii=False,
        )
    )


def _submit_login(args: argparse.Namespace) -> None:
    """フォーム値をstdinからJSON受取して自動送信し、結果をstdoutに出力する。"""
    from dataclasses import asdict

    from crawler.auto_login import submit_login_form

    current_url = getattr(args, "login_current_url", None) or ""
    if not current_url:
        sys.stdout.write(json.dumps({"success": False, "error": "--login-current-url が必要です"}))
        return

    auth_path = Path(args.auth) if args.auth else Path(DEFAULT_AUTH_FILE)
    temp_session = getattr(args, "login_temp_session", None)

    try:
        raw = sys.stdin.read().strip() if not sys.stdin.isatty() else "{}"
        field_values: dict[str, str] = json.loads(raw) if raw else {}
    except json.JSONDecodeError:
        sys.stdout.write(json.dumps({"success": False, "error": "フィールドJSONが不正です"}))
        return

    result = submit_login_form(
        field_values=field_values,
        current_url=current_url,
        auth_path=auth_path,
        temp_session_path=temp_session,
    )
    sys.stdout.write(
        json.dumps(
            {
                "success": result.success,
                "needs_more_fields": result.needs_more_fields,
                "fields": [asdict(f) for f in result.fields],
                "current_url": result.current_url,
                "error": result.error,
            },
            ensure_ascii=False,
        )
    )


def _capture_login(login_url: str, auth_path: Path | None, signal_path: Path | None = None) -> None:
    output_path = Path(auth_path) if auth_path else Path(DEFAULT_AUTH_FILE)
    if signal_path is not None:
        from crawler.auth import capture_auth_state_via_signal

        output_path.parent.mkdir(parents=True, exist_ok=True)
        saved = capture_auth_state_via_signal(login_url, output_path, Path(signal_path))
        if saved is None:
            logger.error("ログイン完了シグナルを待機中にタイムアウトしました")
            sys.exit(1)
        logger.info("ログインセッションを保存しました: %s", saved)
        return
    saved = capture_auth_state(login_url, output_path)
    logger.info(
        "ログインセッションを保存しました: %s （--auth %s でクロールに利用できます）", saved, saved
    )


def _record_login(args: argparse.Namespace) -> None:
    """認証フローレコーダー起動（--login-record、SPEC-3-2）。
    Web UI からの「ブラウザでログインして保存」フローの実処理。"""
    from crawler.auth_recorder import record_auth_session

    login_url = str(getattr(args, "login_record_url", None) or "")
    if not login_url:
        logger.error("--login-record には --login-record-url が必要です")
        sys.exit(1)
    signal_path = getattr(args, "login_signal", None)
    if signal_path is None:
        logger.error("--login-record には --login-signal が必要です")
        sys.exit(1)

    auth_path = getattr(args, "auth", None)
    output_path = Path(auth_path) if auth_path else Path(DEFAULT_AUTH_FILE)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    status_path = getattr(args, "login_status", None)
    # CI・テスト用（実機の目視確認は DoD で別途行う。DISPLAY 前提の headful が既定）
    headless = os.environ.get("WEBSPEC2DOC_RECORD_HEADLESS", "") == "1"

    status = record_auth_session(
        login_url,
        output_path,
        Path(signal_path),
        status_file=Path(status_path) if status_path else None,
        headless=headless,
    )
    logger.info("認証フローレコーダーが終了しました: phase=%s", status.phase)


def _discover(args: argparse.Namespace, auth_path: Path | None) -> None:
    """画面リストを探索し、JSON を stdout に出力する（GUI の画面リスト取得用）。"""
    if not args.url:
        sys.stdout.write(json.dumps({"pages": [], "error": "--url が必要です"}))
        sys.stdout.flush()
        return
    if bool(getattr(args, "stream", False)):

        def _emit(page: dict[str, object]) -> None:
            sys.stdout.write(json.dumps({"page": page}, ensure_ascii=False) + "\n")
            sys.stdout.flush()

        def _emit_discover_event(event: dict[str, object]) -> None:
            sys.stdout.write(json.dumps({"crawl_event": event}, ensure_ascii=False) + "\n")
            sys.stdout.flush()

        pages = discover_pages(
            url=str(args.url),
            depth=int(args.depth),
            max_pages=int(args.max_pages),
            auth_state=Path(auth_path) if auth_path else None,
            on_page_found=_emit,
            on_event=_emit_discover_event,
        )
        sys.stdout.write(json.dumps({"done": True, "total": len(pages)}, ensure_ascii=False) + "\n")
        sys.stdout.flush()
    else:
        pages = discover_pages(
            url=str(args.url),
            depth=int(args.depth),
            max_pages=int(args.max_pages),
            auth_state=Path(auth_path) if auth_path else None,
        )
        sys.stdout.write(json.dumps({"pages": pages}, ensure_ascii=False))


def _compute_transition_quality(
    pages: list[PageData],
) -> tuple[dict[str, dict], list[dict]]:
    """遷移テストパスの N-switch カバレッジとビジネスフロー優先度を算出する。"""
    from graph.transition_graph import (
        business_flows_to_dict,
        classify_pages_for_flows,
        compute_switch_coverage,
        generate_transition_tests,
        prioritize_business_flows,
        switch_coverage_to_dict,
    )

    paths = generate_transition_tests(pages, coverage="1-switch")
    coverage = switch_coverage_to_dict(compute_switch_coverage(pages, paths))
    flows = business_flows_to_dict(
        prioritize_business_flows(paths, classify_pages_for_flows(pages))
    )
    return coverage, flows


def _load_test_metadata(output_dir: Path) -> list[dict]:
    """spec_ts_generator が併産したテストメタデータ JSON（tests リスト）を探索して返す。"""
    candidates = sorted(output_dir.glob("qa_process/*.meta.json")) + sorted(
        output_dir.glob("*.meta.json")
    )
    for meta_path in candidates:
        try:
            data = json.loads(meta_path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError) as exc:
            logger.warning("テストメタデータの読み込みに失敗しました: %s (%s)", meta_path, exc)
            continue
        tests = data.get("tests")
        if isinstance(tests, list):
            return tests
    return []


def _compute_impact_report(
    prior_snapshot: Path,
    analyzed_pages: list[AnalyzedPage],
    output_dir: Path,
) -> dict | None:
    """差分検出→影響テスト特定→再実行推奨リストを report.html 統合表示用に算出する。"""
    from diff.impact_analyzer import (
        analyze_impact,
        build_url_fingerprints,
        format_impact_report,
    )

    try:
        old_pages = load_snapshot(prior_snapshot)
    except (OSError, json.JSONDecodeError) as exc:
        logger.warning("前回スナップショットの読み込みに失敗しました: %s", exc)
        return None
    diff = compute_diff(old_pages, [p.page_data for p in analyzed_pages])
    test_metadata = _load_test_metadata(output_dir)
    url_fingerprints = build_url_fingerprints(analyze_pages(old_pages))
    url_fingerprints.update(build_url_fingerprints(analyzed_pages))
    impacted = analyze_impact(diff, test_metadata, url_fingerprints)
    return format_impact_report(impacted)


def _save_diff_report(
    prior_snapshot: Path | None,
    new_snapshot: Path,
    pages: list[PageData],
    output_dir: Path,
    target_url: str,
) -> bool:
    """差分レポートを生成して保存する。差分がある場合は True を返す。"""
    if prior_snapshot is None:
        logger.info(FIRST_SNAPSHOT_MESSAGE)
        return False
    old_pages = load_snapshot(prior_snapshot)
    diff = compute_diff(old_pages, pages)
    report_html = generate_diff_report(
        diff=diff,
        old_label=prior_snapshot.name,
        new_label=new_snapshot.name,
        target_url=target_url,
    )
    (output_dir / DIFF_REPORT_FILE_NAME).write_text(report_html, encoding="utf-8")
    return bool(getattr(diff, "has_changes", False))


def _run_doc_fusion(
    pages: list[AnalyzedPage],
    reference_docs: list[Path] | None,
    output_dir: Path,
    use_llm: bool = False,
    refresh_doc: bool = False,
) -> tuple[dict[str, str] | None, dict[tuple[str, str], tuple] | None]:
    """参考文書があれば実測結果と突合し、(正式画面名マップ, 文書由来ルール条件) を返す。

    突合結果は doc_fusion.json / doc_fusion.md として出力する。
    文書の取り込み失敗はクロール成果を無駄にしないため警告に留める。
    use_llm=True かつ OPENAI_API_KEY 未設定の場合は Phase 1 抽出のみで継続する。
    refresh_doc=True の場合、突合結果を骨格とした再生版仕様書
    （refreshed_spec.md / refresh_log.json）も併せて出力する。
    """
    if not reference_docs:
        if refresh_doc:
            logger.warning("--refresh-doc は --reference-doc と併用が必須のため無視します")
        return None, None
    from analyzer.rule_injector import build_rule_conditions
    from generator.fusion_reporter import save_fusion_outputs
    from generator.refresh_reporter import save_refresh_outputs
    from generator.trace_reporter import save_trace_outputs
    from ingest.loader import load_reference_documents
    from ingest.matcher import fuse
    from ingest.req_tracer import trace_requirements

    api_key = os.environ.get("OPENAI_API_KEY", "") if use_llm else ""
    try:
        bundle = load_reference_documents(list(reference_docs), use_llm=use_llm, api_key=api_key)
    except (FileNotFoundError, ValueError) as exc:
        logger.warning("参考文書の取り込みに失敗しました（突合をスキップ）: %s", exc)
        return None, None
    result = fuse(pages, bundle)
    save_fusion_outputs(result, bundle, output_dir)
    logger.info(
        "文書×実測の突合が完了しました: 画面対応 %d 件・ギャップ %d 件（doc_fusion.md）",
        len(result.screen_matches),
        len(result.field_gaps),
    )
    if bundle.requirements:
        candidates = _load_playwright_candidates(output_dir)
        traces = trace_requirements(bundle, result, pages, candidates)
        save_trace_outputs(traces, bundle, output_dir)
        logger.info(
            "RFP要件トレーサビリティを出力しました: 要件 %d 件（traceability_matrix.md）",
            len(traces),
        )
    if refresh_doc:
        save_refresh_outputs(result, bundle, pages, output_dir)
        logger.info("文書の再生が完了しました（refreshed_spec.md）")
    rule_conditions = build_rule_conditions(result, bundle, pages) if bundle.rules else None
    return result.official_names or None, rule_conditions or None


def _load_playwright_candidates(output_dir: Path) -> list[dict]:
    """playwright_candidates.json を寛容に読み込む（不在/破損は空リストで継続）。

    web/routes/traceability.py::_load_json_file と同じ寛容な読み方。
    AutoRun 実行前は存在しないため、無いことを異常にしない（candidate_ids=() で継続）。
    """
    path = output_dir / "playwright_candidates.json"
    if not path.exists():
        logger.warning("テスト候補ファイルなし（条件件数のみで追跡）: %s", path)
        return []
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        logger.warning("テスト候補ファイルなし（条件件数のみで追跡）: %s", exc)
        return []
    if isinstance(data, dict):
        return list(data.get("candidates", []))
    if isinstance(data, list):
        return list(data)
    return []


def _build_and_save_ux_review(
    pages: list[PageData],
    analyzed_pages: list[AnalyzedPage],
    ux_axe_results: dict[str, tuple[AxeViolation, ...]],
    output_dir: Path,
) -> dict[str, object]:
    """画面ごとに UX 所見（axe 違反＋ニールセン10原則）を生成し ux_review.json を保存する。

    OPENAI_API_KEY が設定されていれば OpenAIProvider、無ければ RulesProvider で
    完走する（AC-5）。axe 検査結果はクロール中に収集済みの ux_axe_results を用いる
    （axe はライブページが必要なため crawl_page 内でのみ実行できる）。
    """
    from generator.ux_reporter import build_ux_review, build_ux_screen_info, save_ux_outputs
    from llm.provider import OpenAIProvider, RulesProvider

    api_key = os.environ.get("OPENAI_API_KEY", "")
    provider = OpenAIProvider(api_key) if api_key else RulesProvider()

    page_ids = {page.page_data.url: page.page_id for page in analyzed_pages}
    ux_findings: dict[str, list[dict[str, object]]] = {}
    for page in pages:
        axe_violations = ux_axe_results.get(page.url, ())
        screen_info = build_ux_screen_info(page, axe_violations)
        try:
            ux_findings[page.url] = provider.generate_ux_review(screen_info)
        except Exception as exc:  # noqa: BLE001 - UX所見生成の失敗でクロール成果を無駄にしない
            logger.warning("UX 所見の生成に失敗しました（%s）: %s", page.url, exc)
            ux_findings[page.url] = []

    ux_review = build_ux_review(pages, page_ids, ux_axe_results, ux_findings)
    save_ux_outputs(ux_review, output_dir)
    return ux_review


def _load_exploration_coverage(output_dir: Path) -> dict[str, object] | None:
    """既存の exploration_coverage.json を読み込む（無ければ None）。

    破損 JSON は report.html を従来出力へフォールバックさせるため、
    警告に留めて None を返す。
    """
    path = output_dir / "exploration_coverage.json"
    if not path.exists():
        return None
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        logger.warning("探索カバレッジの読込に失敗しました（セクション省略）: %s", exc)
        return None


def _collect_coverage_gaps(
    output_dir: Path,
    pages: list[PageData],
    exploration_coverage: dict[str, object] | None,
) -> tuple:
    """report.html の「カバレッジと未確認領域」節向けにギャップを集計する（AC-5）。

    集計自体の失敗はレポート生成全体を止めない（既存出力へのフォールバック方針は
    _load_exploration_coverage と同様）。
    """
    from generator.coverage_gap import collect_coverage_gaps

    try:
        return collect_coverage_gaps(output_dir, pages, exploration_coverage)
    except Exception as exc:  # noqa: BLE001 - ギャップ集計失敗でレポート生成を止めない
        logger.warning("カバレッジギャップの集計に失敗しました（セクション省略）: %s", exc)
        return ()


def save_outputs(
    pages: list[AnalyzedPage],
    graph: nx.DiGraph,
    form_summary: list[dict[str, object]],
    output_dir: Path,
    formats: Sequence[str],
    llm_insights: object | None = None,
    crawl_depth: int = DEFAULT_DEPTH,
    crawl_max_pages: int = DEFAULT_MAX_PAGES,
    crawled_at: str = "",
    transition_coverage: dict[str, dict] | None = None,
    business_flows: list[dict] | None = None,
    impact_report: dict | None = None,
    official_names: dict[str, str] | None = None,
    exploration_coverage: dict[str, object] | None = None,
    rule_conditions: dict[tuple[str, str], tuple] | None = None,
    ux_review: dict[str, object] | None = None,
    coverage_gaps: tuple = (),
) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    target_url = pages[0].page_data.url if pages else ""
    transition_mmd = generate_mermaid(graph, pages)
    _save_markdown_outputs(pages, graph, form_summary, output_dir, target_url, transition_mmd)
    if "html" in formats or "pdf" in formats:
        _save_html_outputs(
            pages,
            graph,
            form_summary,
            output_dir,
            target_url,
            transition_mmd,
            formats,
            crawl_depth,
            crawl_max_pages,
            crawled_at,
            transition_coverage=transition_coverage,
            business_flows=business_flows,
            impact_report=impact_report,
            exploration_coverage=exploration_coverage,
            ux_review=ux_review,
            coverage_gaps=coverage_gaps,
        )
    if "json" in formats:
        _save_json_output(
            pages,
            graph,
            output_dir,
            target_url,
            crawl_depth,
            crawl_max_pages,
            crawled_at,
            transition_coverage=transition_coverage,
            business_flows=business_flows,
            official_names=official_names,
            rule_conditions=rule_conditions,
        )
    if "excel" in formats:
        _save_excel_output(output_dir, pages, form_summary, official_names)
    if llm_insights is not None:
        logger.warning("llm_insights は現在保存対象外です")


def _save_markdown_outputs(
    pages: list[AnalyzedPage],
    graph: nx.DiGraph,
    form_summary: list[dict[str, object]],
    output_dir: Path,
    target_url: str,
    transition_mmd: str,
) -> None:
    from generator.architecture_generator import (
        generate_architecture_mermaid,
        merge_api_endpoints,
        merge_stack_infos,
    )

    screens_md = generate_screens_markdown(pages, graph, target_url)
    forms_md = generate_forms_markdown(form_summary)
    (output_dir / "screens.md").write_text(screens_md, encoding="utf-8")
    (output_dir / "forms.md").write_text(forms_md, encoding="utf-8")
    (output_dir / "transition.mmd").write_text(transition_mmd, encoding="utf-8")

    stacks = [p.page_data.stack_info for p in pages if p.page_data.stack_info]
    merged_endpoints = merge_api_endpoints([p.page_data.api_calls for p in pages])
    arch_mmd = generate_architecture_mermaid(
        _domain_name(target_url), merge_stack_infos(stacks), merged_endpoints
    )
    (output_dir / "architecture.mmd").write_text(arch_mmd, encoding="utf-8")


def _save_html_outputs(
    pages: list[AnalyzedPage],
    graph: nx.DiGraph,
    form_summary: list[dict[str, object]],
    output_dir: Path,
    target_url: str,
    transition_mmd: str,
    formats: Sequence[str],
    crawl_depth: int,
    crawl_max_pages: int,
    crawled_at: str,
    transition_coverage: dict[str, dict] | None = None,
    business_flows: list[dict] | None = None,
    impact_report: dict | None = None,
    exploration_coverage: dict[str, object] | None = None,
    ux_review: dict[str, object] | None = None,
    coverage_gaps: tuple = (),
) -> None:
    from generator.html_reporter import generate_html_report

    screenshots_dir = output_dir / "screenshots"
    report_html = generate_html_report(
        pages,
        graph,
        form_summary,
        target_url,
        transition_mmd,
        screenshots_dir=screenshots_dir if screenshots_dir.is_dir() else None,
        crawl_depth=crawl_depth,
        crawl_max_pages=crawl_max_pages,
        crawled_at=crawled_at,
        transition_coverage=transition_coverage,
        business_flows=business_flows,
        impact_report=impact_report,
        exploration_coverage=exploration_coverage,
        ux_review=ux_review,
        coverage_gaps=coverage_gaps,
    )
    html_path = output_dir / "report.html"
    html_path.write_text(report_html, encoding="utf-8")
    if "pdf" in formats:
        from generator.pdf_reporter import generate_pdf

        generate_pdf(html_path, output_dir / PDF_FILE_NAME)


def _save_json_output(
    pages: list[AnalyzedPage],
    graph: nx.DiGraph,
    output_dir: Path,
    target_url: str,
    crawl_depth: int,
    crawl_max_pages: int,
    crawled_at: str,
    transition_coverage: dict[str, dict] | None = None,
    business_flows: list[dict] | None = None,
    official_names: dict[str, str] | None = None,
    rule_conditions: dict[tuple[str, str], tuple] | None = None,
) -> None:
    from generator.json_reporter import generate_json_report

    report_json = generate_json_report(
        pages,
        graph,
        target_url,
        crawl_depth=crawl_depth,
        crawl_max_pages=crawl_max_pages,
        crawled_at=crawled_at,
        transition_coverage=transition_coverage,
        business_flows=business_flows,
        official_names=official_names,
        rule_conditions=rule_conditions,
    )
    (output_dir / JSON_REPORT_FILE_NAME).write_text(report_json, encoding="utf-8")


def _parse_formats(raw_formats: str) -> tuple[str, ...]:
    formats = tuple(
        item.strip().lower() for item in raw_formats.split(FORMAT_SEPARATOR) if item.strip()
    )
    unknown = sorted(set(formats) - SUPPORTED_FORMATS)
    if unknown:
        logger.warning("未対応の出力形式を無視します: %s", ", ".join(unknown))
    return tuple(item for item in formats if item in SUPPORTED_FORMATS)


def _domain_name(url: str) -> str:
    parsed = urlparse(url)
    return parsed.netloc or parsed.path.replace("/", "_") or "site"


def _save_excel_output(
    output_dir: Path,
    pages: list[AnalyzedPage],
    form_summary: list[dict[str, object]],
    official_names: dict[str, str] | None = None,
) -> None:
    wb = openpyxl.Workbook()

    _write_screens_sheet(wb.active, pages)
    wb.active.title = "Screens"

    forms_sheet = wb.create_sheet("Forms")
    _write_forms_sheet(forms_sheet, form_summary)

    field_def_sheet = wb.create_sheet("項目定義書")
    _write_field_definitions_sheet(field_def_sheet, pages, official_names)

    bva_sheet = wb.create_sheet("境界値データ")
    _write_bva_sheet(bva_sheet, pages)

    wb.save(output_dir / XLSX_FILE_NAME)


def _write_screens_sheet(
    ws: openpyxl.worksheet.worksheet.Worksheet, pages: list[AnalyzedPage]
) -> None:
    ws.append(["画面ID", "URL", "タイトル", "フォーム数"])
    for page in pages:
        ws.append(
            [page.page_id, page.page_data.url, page.page_data.title, len(page.page_data.forms)]
        )


def _write_forms_sheet(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    form_summary: list[dict[str, object]],
) -> None:
    ws.append(["画面ID", "URL", "フィールド名", "型", "必須", "placeholder", "根拠", "確信度"])
    for item in form_summary:
        ws.append(
            [
                item.get("page_id", ""),
                item.get("url", ""),
                item.get("name", ""),
                item.get("field_type", ""),
                item.get("required", False),
                item.get("placeholder", ""),
                _evidence_cell(item.get("evidence")),
                item.get("confidence", ""),
            ]
        )


def _write_field_definitions_sheet(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    pages: list[AnalyzedPage],
    official_names: dict[str, str] | None = None,
) -> None:
    """実測フィールド属性を SIer 標準の「項目定義書」形式で出力する。"""
    ws.append(
        [
            "画面名",
            "画面ID",
            "URL",
            "項目名",
            "ラベル",
            "型",
            "必須",
            "最小桁",
            "最大桁",
            "範囲",
            "入力規則",
            "選択肢",
            "初期値",
            "placeholder",
            "根拠",
            "確信度",
        ]
    )
    names = official_names or {}
    for page in pages:
        screen_name = names.get(page.page_id) or page.page_data.title
        for form in page.page_data.forms:
            for field in form.fields:
                range_text = (
                    f"{field.min_value}〜{field.max_value}"
                    if (field.min_value or field.max_value)
                    else ""
                )
                ws.append(
                    [
                        screen_name,
                        page.page_id,
                        page.page_data.url,
                        field.name,
                        field.aria_label or "未確認",
                        field.field_type,
                        field.required,
                        field.minlength if field.minlength is not None else "",
                        field.maxlength if field.maxlength is not None else "",
                        range_text,
                        field.pattern,
                        "、".join(field.options),
                        field.default,
                        field.placeholder,
                        _evidence_cell(evidence_to_dict(field.evidence)),
                        field.confidence,
                    ]
                )


def _write_bva_sheet(ws: openpyxl.worksheet.worksheet.Worksheet, pages: list[AnalyzedPage]) -> None:
    """実測属性から機械導出した境界値データを出力する。"""
    from analyzer.bva import KIND_LABELS, attach_observed_boundary_cases, derive_boundary_cases

    ws.append(
        ["画面ID", "項目名", "観点", "入力値", "期待結果", "根拠属性", "根拠セレクタ", "確信度"]
    )
    for page in pages:
        observations = list(page.page_data.validation_observations)
        for form in page.page_data.forms:
            for field in form.fields:
                cases = attach_observed_boundary_cases(
                    derive_boundary_cases(field), field, observations
                )
                for case in cases:
                    value = case.value if case.generated else "（例生成不能 — 手動作成要）"
                    ws.append(
                        [
                            page.page_id,
                            case.field_name,
                            KIND_LABELS.get(case.kind, case.kind),
                            value,
                            case.expected,
                            case.source_attribute,
                            _evidence_cell(evidence_to_dict(case.evidence)),
                            case.confidence,
                        ]
                    )


def _evidence_cell(evidence: object) -> str:
    """evidence dict を Excel セル向けの文字列（セレクタ + 属性）に変換する。"""
    if not isinstance(evidence, dict):
        return ""
    selector = str(evidence.get("selector") or "")
    attribute = evidence.get("html_attribute")
    if attribute:
        return f"{selector} ({attribute})"
    return selector


if __name__ == "__main__":
    main()
