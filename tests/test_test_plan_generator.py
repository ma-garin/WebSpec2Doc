"""テスト計画ドラフト生成（test_plan_generator）のユニット・結合テスト。"""

from __future__ import annotations

import copy
import json
import logging
from pathlib import Path

import openpyxl
import pytest

from generator.test_plan_generator import (
    DISCLAIMER,
    MD_FILE_NAME,
    XLSX_FILE_NAME,
    PlanCoefficients,
    compute_test_plan,
    load_plan_coefficients,
    save_test_plan,
)


def _field(name: str, condition_count: int) -> dict:
    return {
        "name": name,
        "field_type": "text",
        "required": False,
        "test_conditions": [f"cond-{i}" for i in range(condition_count)],
    }


def _screen(
    page_id: str,
    url: str,
    title: str,
    *,
    headings: list[str] | None = None,
    fields: list[dict] | None = None,
    is_canonical: bool = True,
) -> dict:
    forms = [{"action": "", "method": "post", "fields": fields}] if fields else []
    return {
        "page_id": page_id,
        "url": url,
        "title": title,
        "headings": headings or [],
        "forms": forms,
        "is_canonical": is_canonical,
    }


def _report(screens: list[dict], business_flows: list[dict] | None = None) -> dict:
    return {
        "meta": {"business_flows": business_flows or []},
        "screens": screens,
    }


# ---------- 単体テスト ----------


class TestComputeTestPlanRows:
    def test_plan_rows_from_canonical_screens(self) -> None:
        """canonical でない画面（変種）はスコープ表から除外される。"""
        screens = [
            _screen("P001", "https://x/a", "トップ"),
            _screen("P002", "https://x/b", "商品一覧"),
            _screen("P003", "https://x/b?sort=asc", "商品一覧(変種)", is_canonical=False),
        ]
        plan = compute_test_plan(_report(screens), PlanCoefficients())
        assert len(plan.rows) == 2
        page_ids = {row.page_id for row in plan.rows}
        assert page_ids == {"P001", "P002"}
        row = plan.rows[0]
        assert row.page_id
        assert row.title
        assert row.url
        assert row.screen_type
        assert row.test_priority
        assert row.priority_source
        assert isinstance(row.condition_count, int)
        assert isinstance(row.estimated_minutes, float)


class TestEstimateFormula:
    def test_estimate_formula_and_disclaimer(self) -> None:
        """見積 = weight(priority) * minutes_per_screen + condition_count * minutes_per_condition。

        タイトル・見出し・フィールド名は screen_classifier の critical キーワード
        （決済/個人情報/ログイン等）に一致させず、2 フィールド以上で判定される
        "form"（priority=high）に分類されるようにする。
        """
        fields = [_field("name", 2), _field("message", 2)]
        screens = [
            _screen(
                "P005",
                "https://shop.example.com/contact",
                "お問い合わせ",
                headings=["お問い合わせフォーム"],
                fields=fields,
            )
        ]
        plan = compute_test_plan(_report(screens), PlanCoefficients())
        row = plan.rows[0]
        assert row.test_priority == "high"
        assert row.condition_count == 4
        assert row.estimated_minutes == pytest.approx(45.0 * 1.2 + 4 * 10.0)
        assert "係数に基づく推定値であり実測ではない" in plan.disclaimer
        assert plan.disclaimer == DISCLAIMER


class TestBusinessFlowPriorityBoost:
    def test_business_flow_raises_priority(self) -> None:
        """business_flows.nodes に該当する URL は high 未満なら high へ引き上げられる。"""
        screens = [
            _screen(
                "P010",
                "https://x/dashboard",
                "ダッシュボード",
                headings=["ダッシュボード"],
            ),
        ]
        business_flows = [
            {
                "flow_name": "ログイン→決済",
                "path_id": "TP001",
                "nodes": ["https://x/dashboard", "https://x/checkout"],
                "screen_types": ["dashboard", "payment"],
                "priority": "高",
            }
        ]
        plan = compute_test_plan(_report(screens, business_flows), PlanCoefficients())
        row = plan.rows[0]
        # 分類はキーワード非一致で low（一般画面）だったはずが、ビジネスフロー一致で high へ昇格する。
        assert row.test_priority == "high"
        assert "ログイン→決済" in row.priority_source

    def test_flow_does_not_downgrade_already_higher_priority(self) -> None:
        """既に critical/high の画面はビジネスフロー一致で降格されない。"""
        fields = [_field("password", 1)]
        screens = [
            _screen(
                "P011",
                "https://x/login",
                "ログイン",
                headings=["ログイン"],
                fields=fields,
            ),
        ]
        business_flows = [
            {
                "flow_name": "ログイン",
                "path_id": "TP002",
                "nodes": ["https://x/login"],
                "screen_types": ["login"],
                "priority": "高",
            }
        ]
        plan = compute_test_plan(_report(screens, business_flows), PlanCoefficients())
        row = plan.rows[0]
        assert row.test_priority in {"critical", "high"}


class TestEnvOverride:
    def test_env_override_and_invalid_fallback(
        self, monkeypatch: pytest.MonkeyPatch, caplog: pytest.LogCaptureFixture
    ) -> None:
        monkeypatch.setenv("WEBSPEC2DOC_PLAN_WEIGHT_HIGH", "2.0")
        coefficients = load_plan_coefficients()
        assert coefficients.weight_high == 2.0

        monkeypatch.setenv("WEBSPEC2DOC_PLAN_WEIGHT_HIGH", "abc")
        with caplog.at_level(logging.WARNING):
            coefficients = load_plan_coefficients()
        assert coefficients.weight_high == 1.2
        assert "WEBSPEC2DOC_PLAN_WEIGHT_HIGH" in caplog.text


class TestMissingBusinessFlowsKey:
    def test_missing_business_flows_key(self) -> None:
        """meta に business_flows が無くても例外にならず分類優先度のみで生成される。"""
        screens = [_screen("P001", "https://x/a", "トップ")]
        report = {"meta": {}, "screens": screens}
        plan = compute_test_plan(report, PlanCoefficients())
        assert len(plan.rows) == 1
        assert plan.rows[0].priority_source == "画面分類"


class TestEmptyScreensPlan:
    def test_empty_screens_plan(self) -> None:
        plan = compute_test_plan(_report([]), PlanCoefficients())
        assert plan.rows == ()
        assert plan.total_minutes == 0
        assert plan.total_hours == 0


class TestCoefficientsParity:
    def test_coefficients_parity_with_usage_tracker(self) -> None:
        """AC-7: src 側の既定値が web/services/usage_tracker.py の既定値と一致する。"""
        from web.services.usage_tracker import (
            MINUTES_PER_SCREEN_SPEC,
            MINUTES_PER_TEST_CONDITION,
        )

        from generator.test_plan_generator import (
            MINUTES_PER_CONDITION_DEFAULT,
            MINUTES_PER_SCREEN_DEFAULT,
        )

        assert MINUTES_PER_SCREEN_DEFAULT == MINUTES_PER_SCREEN_SPEC
        assert MINUTES_PER_CONDITION_DEFAULT == MINUTES_PER_TEST_CONDITION


# ---------- 結合テスト（実ファイル I/O） ----------


class TestSaveOutputs:
    def test_save_outputs_md_and_xlsx(self, tmp_path: Path) -> None:
        fields = [_field("card_number", 3)]
        screens = [
            _screen(
                "P004",
                "https://x/checkout",
                "お支払い",
                headings=["お支払い"],
                fields=fields,
            ),
            _screen("P001", "https://x/", "トップ"),
        ]
        plan = compute_test_plan(_report(screens), PlanCoefficients())
        save_test_plan(plan, tmp_path)

        md_path = tmp_path / MD_FILE_NAME
        assert md_path.exists()
        md_text = md_path.read_text(encoding="utf-8")
        assert "スコープ表" in md_text
        assert "見積サマリ" in md_text
        assert DISCLAIMER in md_text

        xlsx_path = tmp_path / XLSX_FILE_NAME
        assert xlsx_path.exists()
        wb = openpyxl.load_workbook(xlsx_path)
        assert wb.sheetnames == ["スコープ表", "見積サマリ"]
        scope_sheet = wb["スコープ表"]
        # ヘッダ行 + 2 画面分
        assert scope_sheet.max_row == 3

    def test_empty_screens_writes_note_in_markdown(self, tmp_path: Path) -> None:
        plan = compute_test_plan(_report([]), PlanCoefficients())
        save_test_plan(plan, tmp_path)
        md_text = (tmp_path / MD_FILE_NAME).read_text(encoding="utf-8")
        assert "対象画面 0 件" in md_text

    def test_xlsx_failure_still_writes_md(
        self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch, caplog: pytest.LogCaptureFixture
    ) -> None:
        """xlsx 書き込みが失敗しても md は出力され、警告ログが出る。"""
        import generator.test_plan_generator as mod

        def _boom(_plan: object, _output_dir: Path) -> None:
            raise OSError("disk full")

        monkeypatch.setattr(mod, "_write_xlsx", _boom)
        plan = compute_test_plan(
            _report([_screen("P001", "https://x/", "トップ")]), PlanCoefficients()
        )
        with caplog.at_level(logging.WARNING):
            save_test_plan(plan, tmp_path)
        assert (tmp_path / MD_FILE_NAME).exists()
        assert not (tmp_path / XLSX_FILE_NAME).exists()
        assert "test_plan.xlsx" in caplog.text


class TestMainCliIntegration:
    def test_missing_report_logs_error(
        self, tmp_path: Path, caplog: pytest.LogCaptureFixture
    ) -> None:
        import argparse

        from main import _generate_test_plan

        args = argparse.Namespace(url="https://x.example.com/", output=tmp_path)
        with caplog.at_level(logging.ERROR):
            _generate_test_plan(args)
        assert "クロール済みインベントリがありません" in caplog.text
        assert not (tmp_path / "x-example-com" / MD_FILE_NAME).exists()

    def test_corrupt_report_logs_error(
        self, tmp_path: Path, caplog: pytest.LogCaptureFixture
    ) -> None:
        import argparse

        from main import _domain_name, _generate_test_plan

        url = "https://x.example.com/"
        output_dir = tmp_path / _domain_name(url)
        output_dir.mkdir(parents=True)
        (output_dir / "report.json").write_text("{not valid json", encoding="utf-8")

        args = argparse.Namespace(url=url, output=tmp_path)
        with caplog.at_level(logging.ERROR):
            _generate_test_plan(args)
        assert "report.json を読み込めません" in caplog.text

    def test_cli_generates_test_plan_from_report(self, tmp_path: Path) -> None:
        import argparse

        from main import _domain_name, _generate_test_plan

        url = "https://x.example.com/"
        output_dir = tmp_path / _domain_name(url)
        output_dir.mkdir(parents=True)
        report = _report(
            [
                _screen(
                    "P004",
                    "https://x.example.com/checkout",
                    "お支払い",
                    headings=["お支払い"],
                    fields=[_field("card_number", 2)],
                )
            ]
        )
        (output_dir / "report.json").write_text(
            json.dumps(report, ensure_ascii=False), encoding="utf-8"
        )

        args = argparse.Namespace(url=url, output=tmp_path)
        _generate_test_plan(args)

        assert (output_dir / MD_FILE_NAME).exists()
        assert (output_dir / XLSX_FILE_NAME).exists()

    def test_report_untouched_by_test_plan_generation(self, tmp_path: Path) -> None:
        """AC-6: report.json は読み取り専用で扱われ、内容が変化しない。"""
        import argparse

        from main import _domain_name, _generate_test_plan

        url = "https://x.example.com/"
        output_dir = tmp_path / _domain_name(url)
        output_dir.mkdir(parents=True)
        report = _report([_screen("P001", "https://x.example.com/", "トップ")])
        report_text = json.dumps(report, ensure_ascii=False)
        (output_dir / "report.json").write_text(report_text, encoding="utf-8")
        original = copy.deepcopy(report)

        args = argparse.Namespace(url=url, output=tmp_path)
        _generate_test_plan(args)

        after_text = (output_dir / "report.json").read_text(encoding="utf-8")
        assert after_text == report_text
        assert json.loads(after_text) == original
