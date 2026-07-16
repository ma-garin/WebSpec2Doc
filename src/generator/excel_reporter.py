"""実測仕様を Excel ワークブックへ出力する。"""

from __future__ import annotations

from pathlib import Path

import openpyxl

from analyzer.html_analyzer import AnalyzedPage
from crawler.page_crawler import evidence_to_dict

XLSX_FILE_NAME = "spec.xlsx"


def save_excel_output(
    output_dir: Path,
    pages: list[AnalyzedPage],
    form_summary: list[dict[str, object]],
    official_names: dict[str, str] | None = None,
) -> None:
    wb = openpyxl.Workbook()

    _write_screens_sheet(wb.active, pages)
    wb.active.title = "Screens"

    forms_sheet = wb.create_sheet("Forms")
    _write_forms_sheet(forms_sheet, form_summary)

    field_def_sheet = wb.create_sheet("項目定義書")
    _write_field_definitions_sheet(field_def_sheet, pages, official_names)

    bva_sheet = wb.create_sheet("境界値データ")
    _write_bva_sheet(bva_sheet, pages)

    wb.save(output_dir / XLSX_FILE_NAME)


def _write_screens_sheet(
    ws: openpyxl.worksheet.worksheet.Worksheet, pages: list[AnalyzedPage]
) -> None:
    ws.append(["画面ID", "URL", "タイトル", "フォーム数"])
    for page in pages:
        ws.append(
            [page.page_id, page.page_data.url, page.page_data.title, len(page.page_data.forms)]
        )


def _write_forms_sheet(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    form_summary: list[dict[str, object]],
) -> None:
    ws.append(["画面ID", "URL", "フィールド名", "型", "必須", "placeholder", "根拠", "確信度"])
    for item in form_summary:
        ws.append(
            [
                item.get("page_id", ""),
                item.get("url", ""),
                item.get("name", ""),
                item.get("field_type", ""),
                item.get("required", False),
                item.get("placeholder", ""),
                _evidence_cell(item.get("evidence")),
                item.get("confidence", ""),
            ]
        )


def _write_field_definitions_sheet(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    pages: list[AnalyzedPage],
    official_names: dict[str, str] | None = None,
) -> None:
    """実測フィールド属性を SIer 標準の「項目定義書」形式で出力する。"""
    ws.append(
        [
            "画面名",
            "画面ID",
            "URL",
            "項目名",
            "ラベル",
            "型",
            "必須",
            "最小桁",
            "最大桁",
            "範囲",
            "入力規則",
            "選択肢",
            "初期値",
            "placeholder",
            "根拠",
            "確信度",
        ]
    )
    names = official_names or {}
    for page in pages:
        screen_name = names.get(page.page_id) or page.page_data.title
        for form in page.page_data.forms:
            for field in form.fields:
                range_text = (
                    f"{field.min_value}〜{field.max_value}"
                    if (field.min_value or field.max_value)
                    else ""
                )
                ws.append(
                    [
                        screen_name,
                        page.page_id,
                        page.page_data.url,
                        field.name,
                        field.aria_label or "未確認",
                        field.field_type,
                        field.required,
                        field.minlength if field.minlength is not None else "",
                        field.maxlength if field.maxlength is not None else "",
                        range_text,
                        field.pattern,
                        "、".join(field.options),
                        field.default,
                        field.placeholder,
                        _evidence_cell(evidence_to_dict(field.evidence)),
                        field.confidence,
                    ]
                )


def _write_bva_sheet(ws: openpyxl.worksheet.worksheet.Worksheet, pages: list[AnalyzedPage]) -> None:
    """実測属性から機械導出した境界値データを出力する。"""
    from analyzer.bva import KIND_LABELS, attach_observed_boundary_cases, derive_boundary_cases

    ws.append(
        ["画面ID", "項目名", "観点", "入力値", "期待結果", "根拠属性", "根拠セレクタ", "確信度"]
    )
    for page in pages:
        observations = list(page.page_data.validation_observations)
        for form in page.page_data.forms:
            for field in form.fields:
                cases = attach_observed_boundary_cases(
                    derive_boundary_cases(field), field, observations
                )
                for case in cases:
                    value = case.value if case.generated else "（例生成不能 — 手動作成要）"
                    ws.append(
                        [
                            page.page_id,
                            case.field_name,
                            KIND_LABELS.get(case.kind, case.kind),
                            value,
                            case.expected,
                            case.source_attribute,
                            _evidence_cell(evidence_to_dict(case.evidence)),
                            case.confidence,
                        ]
                    )


def _evidence_cell(evidence: object) -> str:
    """evidence dict を Excel セル向けの文字列へ変換する。"""
    if not isinstance(evidence, dict):
        return ""
    selector = str(evidence.get("selector") or "")
    attribute = evidence.get("html_attribute")
    if attribute:
        return f"{selector} ({attribute})"
    return selector
